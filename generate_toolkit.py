"""
generate_toolkit.py  –  Forus Resilience & Support Toolkit
Increment VERSION below each time you make a significant change.
Run: python generate_toolkit.py
Produces: Forus_Toolkit_v{VERSION}_Public.pdf and Forus_Toolkit_v{VERSION}_Network.pdf
"""

import openpyxl, os, sys, re, datetime, io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate, Paragraph,
    Spacer, Table, TableStyle, KeepTogether, HRFlowable, PageBreak
)
from reportlab.platypus.flowables import Flowable

# ── Version — increment this each build ──────────────────────────────────────
VERSION     = "1.9"
DATE_STAMP  = datetime.date.today().strftime("%Y-%m-%d")
SPREADSHEET = "Forus_Toolkit_Content_DB.xlsx"
OUT_PUBLIC  = f"/mnt/user-data/outputs/Forus_Toolkit_v{VERSION}_Public.pdf"
OUT_NETWORK = f"/mnt/user-data/outputs/Forus_Toolkit_v{VERSION}_Network.pdf"

# ── Colours ───────────────────────────────────────────────────────────────────
def h(hex_):
    r,g,b = int(hex_[0:2],16), int(hex_[2:4],16), int(hex_[4:6],16)
    return colors.Color(r/255, g/255, b/255)

# ── Forus brand palette (Brand Manual September 2024) ──
#   Primary:   Light Blue #58C5C7 · Dark Blue/Teal #00424D
#   Secondary: Hot Pink #ED1651  · Mint Green #5C9C8E · Lime Green #B2C100
C = {
    # Forus primary
    "forus_teal":       h("58C5C7"),   # dominant brand colour – Light Blue
    "forus_dark":       h("00424D"),   # Dark Blue/Teal
    "forus_pink":       h("ED1651"),   # Hot Pink – impact & action
    "forus_mint":       h("5C9C8E"),   # Mint Green
    "forus_lime":       h("B2C100"),   # Lime Green (CMYK C30 M05 Y100 K00)
    # Forus tinted backgrounds
    "forus_teal_lt":    h("D4F0F1"),   # light teal background
    "forus_pink_lt":    h("FCE4EC"),   # light pink background
    "forus_mint_lt":    h("E0F2EE"),   # light mint background
    "forus_lime_lt":    h("F3F7DB"),   # light lime background

    # Semantic aliases that the renderer uses (mapped to Forus equivalents)
    "dark_green":       h("00424D"),   # → Forus Dark
    "mid_green":        h("5C9C8E"),   # → Forus Mint
    "light_green":      h("E0F2EE"),   # → Forus Mint light
    "teal":             h("58C5C7"),   # → Forus Teal
    "light_teal":       h("D4F0F1"),   # → Forus Teal light
    "deep_red":         h("ED1651"),   # → Forus Pink (crisis/warning accent)
    "light_red":        h("FCE4EC"),   # → Forus Pink light
    "blue":             h("00424D"),   # → Forus Dark (link / db-ref colour)
    "light_blue":       h("D4F0F1"),   # → Forus Teal light
    "purple":           h("00424D"),   # → Forus Dark (template header)
    "light_purple":     h("E4EFF0"),   # soft background
    "brown":            h("5C9C8E"),   # → Forus Mint (case-example header)
    "light_brown":      h("E0F2EE"),   # → Forus Mint light
    "amber":            h("B2C100"),   # → Forus Lime (decision / review)
    "light_amber":      h("F3F7DB"),   # → Forus Lime light
    "orange":           h("ED1651"),   # → Forus Pink (gap notes)
    "grey":             h("3A3A3A"),   # near-black text
    "mid_grey":         h("888888"),   # muted text / rules
    "light_grey":       h("F5F5F5"),   # zebra / step backgrounds
    "white":            h("FFFFFF"),
}

# Part colours follow Forus brand palette
PART_COLORS = {
    1: h("ED1651"),   # Hot Pink   — Crisis Guides
    2: h("58C5C7"),   # Teal       — Solidarity Activation
    3: h("00424D"),   # Dark Blue  — Legal Support
    4: h("5C9C8E"),   # Mint Green — Emergency Funding
    5: h("B2C100"),   # Lime Green — Safe Comms
    6: h("00424D"),   # Dark Blue  — Diversification
    7: h("888888"),   # Mid Grey   — Feedback
    8: h("3A3A3A"),   # Near-black — Annexes
}
PART_LABELS = {
    1:"CRISIS GUIDES", 2:"SOLIDARITY", 3:"LEGAL SUPPORT",
    4:"EMERGENCY FUNDING", 5:"SAFE COMMS", 6:"DIVERSIFICATION",
    7:"FEEDBACK", 8:"ANNEXES",
}

# Simplified two-state time horizons (preemptive / responsive)
# Old values (first-hour, first-24hrs, first-72hrs) all map to responsive display
TIME_LABELS = {
    "preemptive":  "🛡  PREEMPTIVE ACTIONS",
    "responsive":  "⚡  RESPONSIVE ACTIONS",
    # Legacy values — rendered as responsive
    "first-hour":  "⚡  RESPONSIVE ACTIONS",
    "first-24hrs": "⚡  RESPONSIVE ACTIONS",
    "first-72hrs": "⚡  RESPONSIVE ACTIONS",
    "general":     "⚡  RESPONSIVE ACTIONS",
}

# ── Layout ────────────────────────────────────────────────────────────────────
PAGE_W, PAGE_H = A4
NAV_W   = 8*mm
ML      = 18*mm          # left margin (nav strip sits here)
MR      = 14*mm
MT      = 18*mm
MB      = 18*mm
FRAME_X = ML + NAV_W + 4*mm
FRAME_W = PAGE_W - FRAME_X - MR

WORD_LIMITS = {
    "STEP":25,"INTRO":60,"TIP":35,"WARNING":35,
    "DECISION-Q":20,"DECISION-A":8,"CHECKLIST":20,
    "TEMPLATE":120,"CASE":60,"HEADER":10,"DB-REF":40,"MECHANISM-REF":15,
    "SUBSECTION":12,"FEEDBACK":120,
    "REGION-NAV":80,          # pipe-separated region names → clickable buttons to annex
    "COUNTRY-ENTRY":80,       # "Country Name | details" → bold header + details in annex
    "PEER-CONNECT":20,        # legacy key (kept for back-compat)
    "PEER-CONNECTION":110,    # org description (v2.2+)
}

# Two-state sort: preemptive first, everything else second
TIME_ORDER = {
    "preemptive": 0,
    "general":    1,
    "responsive": 1,
    "first-hour": 1,
    "first-24hrs":1,
    "first-72hrs":1,
}

# ── Styles ────────────────────────────────────────────────────────────────────
def ps(name, font="Helvetica", size=9, leading=13, color="grey", bold=False,
       italic=False, align=TA_LEFT, before=0, after=0):
    fn = "Helvetica-BoldOblique" if bold and italic else \
         "Helvetica-Bold" if bold else \
         "Helvetica-Oblique" if italic else font
    return ParagraphStyle(name, fontName=fn, fontSize=size, leading=leading,
                          textColor=C[color] if isinstance(color,str) else color,
                          alignment=align, spaceBefore=before, spaceAfter=after)

S = {
    "intro":    ps("intro",   size=9.5, leading=14, color="grey",       before=2, after=4),
    "step":     ps("step",    size=9,   leading=13, color="grey"),
    "tip":      ps("tip",     size=9,   leading=13, color="dark_green"),
    "warn":     ps("warn",    size=9,   leading=13, color="deep_red"),
    "decQ":     ps("decQ",    size=9,   leading=13, color="amber",  bold=True),
    "decA":     ps("decA",    size=8.5, leading=12, color="grey"),
    "check":    ps("check",   size=9,   leading=13, color="grey"),
    "tmpl":     ps("tmpl", font="Courier", size=8, leading=11.5, color="grey"),
    "case":     ps("case",    size=9,   leading=13, color="grey",   italic=True),
    "dbref":    ps("dbref",   size=8.5, leading=12, color="blue"),
    "mechref":  ps("mechref", size=8.5, leading=12, color="dark_green", bold=True),
    "mname":    ps("mname",   size=9.5, leading=13, color="white",  bold=True),
    "mlabel":   ps("mlabel",  size=7,   leading=9,  color="mid_grey",bold=True),
    "mfield":   ps("mfield",  size=8.5, leading=12, color="grey"),
    "tagw":     ps("tagw",    size=7.5, leading=9,  color="white",  bold=True, align=TA_CENTER),
    "tagd":     ps("tagd",    size=7.5, leading=9,  color="grey",   bold=True, align=TA_CENTER),
    "hdr":      ps("hdr",     size=13,  leading=16, color="white",  bold=True),
    "hdr_sub":  ps("hdr_sub", size=7.5, leading=10, color="light_green", bold=True),
    "footer":   ps("footer",  size=7,   leading=9,  color="mid_grey"),
    "cover_t":  ps("cover_t", size=24,  leading=30, color="white",  bold=True, align=TA_CENTER),
    "cover_s":  ps("cover_s", size=11,  leading=15, color="light_green", align=TA_CENTER),
    "cover_m":  ps("cover_m", size=8.5, leading=12, color="light_green", align=TA_CENTER),
    "toc_h":    ps("toc_h",   size=10,  leading=14, color="white",  bold=True),
    "toc_s":    ps("toc_s",   size=8.5, leading=12, color="grey"),
    "over":     ps("over",    size=8,   leading=10, color="deep_red",bold=True),
    "normal":   ps("normal",  size=9,   leading=13, color="grey"),
    "stepnum":  ps("stepnum", size=11,  leading=14, color="white",  bold=True, align=TA_CENTER),
    "timeline_active": ps("tl_a", size=8, leading=10, color="white", bold=True, align=TA_CENTER),
    "timeline_idle":   ps("tl_i", size=8, leading=10, color="mid_grey", bold=True, align=TA_CENTER),
    "section_label": ps("sl",  size=7.5, leading=9,  color="light_green", bold=True),
    "entry_q":  ps("entry_q", size=8.5, leading=12, color="grey",   bold=True),
    "entry_a":  ps("entry_a", size=8.5, leading=12, color="dark_green", bold=True),
    "gap_note": ps("gn",      size=7.5, leading=10, color="orange", italic=True),
    "sub_label": ps("sub_lbl", size=8,   leading=10, color="grey",   bold=True),
    "sub_title": ps("sub_ttl", size=11,  leading=14, color="grey",   bold=True),
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def wc(text):
    t = str(text or "").strip()
    return len(t.split()) if t else 0

def get_limit(item):
    """Per-row word limit: use spreadsheet value if present, else type default.
    Editor can override by typing any number into the word_limit column."""
    row_val = item.get("word_limit")
    if row_val is not None and str(row_val).strip() not in ("", "None"):
        try:
            return int(row_val)
        except (ValueError, TypeError):
            pass
    btype = str(item.get("block_type", "")).strip()
    return WORD_LIMITS.get(btype, 9999)

def trim(text, btype_or_limit):
    """Accept either a block type string (looks up default) or an int limit directly."""
    if isinstance(btype_or_limit, int):
        limit = btype_or_limit
    else:
        limit = WORD_LIMITS.get(btype_or_limit, 9999)
    words = str(text or "").strip().split()
    if len(words) > limit:
        return " ".join(words[:limit]) + " [...]", True
    return " ".join(words), False

def tag_cell(text, bg, fg=C["white"], pad=(3,6,3,6)):
    return (Paragraph(text, ps("_tag", size=7.5, leading=9,
            color=fg if isinstance(fg,str) else fg, bold=True, align=TA_CENTER)),
            bg, pad)

def ts(*args):
    """Shorthand TableStyle constructor from list of commands."""
    return TableStyle(list(args))

def plain_table(data, col_widths, style_cmds, row_heights=None):
    t = Table(data, colWidths=col_widths, rowHeights=row_heights)
    t.setStyle(TableStyle(style_cmds))
    return t

PAD = (("LEFTPADDING",(0,0),(-1,-1),5),
       ("RIGHTPADDING",(0,0),(-1,-1),5),
       ("TOPPADDING",(0,0),(-1,-1),4),
       ("BOTTOMPADDING",(0,0),(-1,-1),4))

# ── Block renderers (all return a list of flowables) ─────────────────────────

def render_feedback(url):
    """Prominent call-to-action box with a clickable URL."""
    label_row = [[
        Paragraph("HELP KEEP THIS TOOLKIT CURRENT", ps(
            "_fblbl", size=8, leading=10, bold=True,
            color=C["white"], align=TA_CENTER)),
    ]]
    label_t = Table(label_row, colWidths=[FRAME_W])
    label_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), C["teal"]),
        ("TOPPADDING", (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING", (0,0),(-1,-1), 8),
    ]))

    body_row = [[
        Paragraph(
            "If you find something out of date, incorrect, or missing, please use the link below "
            "to send feedback to the Forus team. It takes two minutes and helps every platform "
            "that uses this resource.",
            ps("_fbbody", size=9.5, leading=14, color=C["grey"])),
    ]]
    body_t = Table(body_row, colWidths=[FRAME_W])
    body_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), C["light_teal"]),
        ("TOPPADDING", (0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING", (0,0),(-1,-1), 10),
        ("RIGHTPADDING",(0,0),(-1,-1), 10),
    ]))

    link_row = [[
        Paragraph(
            f'<a href="{url}" color="#1F4E79"><u>{url}</u></a>',
            ps("_fburl", size=11, leading=15, bold=True, color=C["blue"])),
    ]]
    link_t = Table(link_row, colWidths=[FRAME_W])
    link_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,-1), C["light_teal"]),
        ("TOPPADDING", (0,0),(-1,-1), 2),
        ("BOTTOMPADDING",(0,0),(-1,-1), 10),
        ("LEFTPADDING", (0,0),(-1,-1), 10),
        ("RIGHTPADDING",(0,0),(-1,-1), 10),
    ]))

    outer = Table([[label_t],[body_t],[link_t]], colWidths=[FRAME_W])
    outer.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 1.5, C["teal"]),
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 0),
        ("TOPPADDING",   (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
        ("ROUNDEDCORNERS",[4,4,4,4]),
    ]))
    return [Spacer(1, 4*mm), outer, Spacer(1, 4*mm)]


def render_section_banner(text, part):
    pc = PART_COLORS.get(part, C["dark_green"])
    pl = PART_LABELS.get(part, f"PART {part}")
    data = [[
        Paragraph(f"PART {part}  ·  {pl}", S["section_label"]),
        Paragraph(str(part), S["stepnum"]),
    ],[
        Paragraph(text, S["hdr"]),
        "",
    ]]
    t = Table(data, colWidths=[FRAME_W - 14*mm, 14*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), pc),
        ("SPAN",(1,0),(1,1)),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("VALIGN",(1,0),(1,1),"MIDDLE"),
        ("ALIGN",(1,0),(1,1),"CENTER"),
        ("TOPPADDING",(0,0),(-1,-1),7),
        ("BOTTOMPADDING",(0,0),(-1,-1),7),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("ROUNDEDCORNERS",[3,3,3,3]),
    ]))
    return [t, Spacer(1, 5*mm)]

def render_subsection_divider(text, part, time_horizon):
    """Labelled divider — more prominent than a timeline bar tab, lighter than a section banner."""
    pc = PART_COLORS.get(part, C["dark_green"])

    # Accent colour varies by horizon so preemptive has a distinct visual identity
    HORIZON_ACCENTS = {
        "preemptive":  h("3D6B8A"),   # steel blue — forward-looking
        "first-hour":  C["deep_red"],
        "first-24hrs": C["amber"],
        "first-72hrs": C["mid_green"],
        "general":     C["grey"],
    }
    accent = HORIZON_ACCENTS.get(time_horizon, pc)

    icon_map = {
        "preemptive":  "🛡",
        "first-hour":  "⏱",
        "first-24hrs": "📋",
        "first-72hrs": "🔄",
        "general":     "·",
    }
    icon = icon_map.get(time_horizon, "·")

    data = [[
        Paragraph(f"{icon}  {text.upper()}", ps(
            "_sub", size=10, leading=13, bold=True,
            color=accent)),
    ]]
    t = Table(data, colWidths=[FRAME_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), C["light_grey"]),
        ("LINEBELOW",     (0,0),(-1,-1), 2, accent),
        ("LINEABOVE",     (0,0),(-1,-1), 0.3, C["mid_grey"]),
        ("LEFTPADDING",   (0,0),(-1,-1), 5),
        ("RIGHTPADDING",  (0,0),(-1,-1), 5),
        ("TOPPADDING",    (0,0),(-1,-1), 5),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
    ]))
    return [Spacer(1, 6*mm), t, Spacer(1, 4*mm)]

def render_timeline_bar(active_horizon):
    """Two-state bar: Preemptive Actions | Responsive Actions."""
    is_preemptive = (active_horizon == "preemptive")
    states = [
        ("preemptive", "🛡  PREEMPTIVE ACTIONS"),
        ("responsive", "⚡  RESPONSIVE ACTIONS"),
    ]
    cells = []
    for key, label in states:
        active = (key == "preemptive") == is_preemptive
        cells.append(Paragraph(label,
            S["timeline_active"] if active else S["timeline_idle"]))
    w = FRAME_W / 2
    t = Table([cells], colWidths=[w, w], rowHeights=[8*mm])
    cmds = [
        ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("BACKGROUND",(0,0),(0,0), C["dark_green"] if is_preemptive else C["light_grey"]),
        ("BACKGROUND",(1,0),(1,0), C["dark_green"] if not is_preemptive else C["light_grey"]),
    ]
    t.setStyle(TableStyle(cmds))
    return [t, Spacer(1,1.5*mm)]

def render_step(text, number, part, truncated):
    pc = PART_COLORS.get(part, C["dark_green"])
    suffix = " <font color='red'>[OVER LIMIT]</font>" if truncated else ""
    # Number cell: stacked — big number on top, small "ACTION STEP" label below
    num_cell = Table([
        [Paragraph(str(number), S["stepnum"])],
        [Paragraph("ACTION<br/>STEP", ps("_stype", size=5.5, leading=7,
                    color=C["white"], bold=True, align=TA_CENTER))],
    ], colWidths=[10*mm])
    num_cell.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), pc),
        ("TOPPADDING",(0,0),(0,0), 6),("BOTTOMPADDING",(0,0),(0,0), 1),
        ("TOPPADDING",(0,1),(0,1), 0),("BOTTOMPADDING",(0,1),(0,1), 5),
        ("LEFTPADDING",(0,0),(-1,-1), 1),("RIGHTPADDING",(0,0),(-1,-1), 1),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))
    data = [[num_cell, Paragraph(_linkify_refs(text) + suffix, S["step"])]]
    t = Table(data, colWidths=[10*mm, FRAME_W - 10*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(1,0),(1,0), C["light_grey"]),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(0,0),0),("RIGHTPADDING",(0,0),(0,0),0),
        ("LEFTPADDING",(1,0),(1,0),8),("RIGHTPADDING",(1,0),(1,0),6),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("LINEAFTER",(0,0),(0,0), 1.5, pc),
    ]))
    return [t, Spacer(1, 2.5*mm)]

def render_callout(text, box_type, truncated):
    if box_type == "TIP":
        bg, accent, icon = C["light_green"], C["mid_green"], "✓  TIP"
        style = S["tip"]
    else:
        bg, accent, icon = C["light_red"], C["deep_red"], "!  WARNING"
        style = S["warn"]
    suffix = " <font color='red'>[OVER LIMIT]</font>" if truncated else ""
    data = [[
        Paragraph(icon, ps("_icon", size=7, leading=9,
                           color=accent, bold=True, align=TA_CENTER)),
        Paragraph(_linkify_refs(text) + suffix, style),
    ]]
    t = Table(data, colWidths=[14*mm, FRAME_W - 14*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0), accent),
        ("BACKGROUND",(1,0),(1,0), bg),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(0,0),2),("RIGHTPADDING",(0,0),(0,0),2),
        ("LEFTPADDING",(1,0),(1,0),8),("RIGHTPADDING",(1,0),(1,0),6),
        ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("ROUNDEDCORNERS",[2,2,2,2]),
    ]))
    return [t, Spacer(1, 2.5*mm)]

def render_intro(text):
    label = Table(
        [[Paragraph("CONTEXT", ps("_il", size=6.5, leading=9,
                                   color=C["mid_grey"], bold=True))]],
        colWidths=[FRAME_W])
    label.setStyle(TableStyle([
        ("LINEBELOW",(0,0),(-1,-1), 0.75, C["forus_teal"]),
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),2),
    ]))
    return [label, Paragraph(_linkify_refs(text), S["intro"]), Spacer(1, 4*mm)]


def render_header(text):
    """Phase/structural header banner (BEFORE CRISIS, DURING CRISIS, etc.)."""
    tl = text.upper()
    if "BEFORE" in tl:
        bg, border = C["dark_green"], C["mid_green"]
        icon = "🛡  "
    elif "DURING" in tl:
        bg, border = C["deep_red"], C["orange"]
        icon = "⚡  "
    else:
        bg, border = C["forus_dark"], C["forus_teal"]
        icon = ""
    t = Table(
        [[Paragraph(icon + text.upper(),
                    ps("_hdr_txt", size=9.5, leading=12,
                       color=C["white"], bold=True, align=TA_CENTER))]],
        colWidths=[FRAME_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), bg),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW",     (0, 0), (-1, -1), 2, border),
    ]))
    return [Spacer(1, 5*mm), t, Spacer(1, 3*mm)]


def render_decision_q(text):
    data = [[Paragraph(f"⬥  DECISION", ps("_dq_lbl",size=7,leading=9,
                       color=C["amber"],bold=True,align=TA_CENTER)),
             Paragraph(text, S["decQ"])]]
    t = Table(data, colWidths=[18*mm, FRAME_W-18*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0), C["light_amber"]),
        ("BACKGROUND",(1,0),(1,0), C["light_amber"]),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LINEBELOW",(0,0),(-1,-1),0.5,C["amber"]),
    ]))
    return [t, Spacer(1,1*mm)]

def render_decision_a(text, truncated):
    suffix = " [OVER LIMIT]" if truncated else ""
    data = [[
        Paragraph("→  OPTION", ps("_da_lbl", size=6.5, leading=9,
                                   color=C["amber"], bold=True, align=TA_CENTER)),
        Paragraph(text + suffix, S["decA"]),
    ]]
    t = Table(data, colWidths=[16*mm, FRAME_W - 16*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0), C["light_amber"]),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(0,0),3),("RIGHTPADDING",(0,0),(0,0),3),
        ("LEFTPADDING",(1,0),(1,0),7),("RIGHTPADDING",(1,0),(1,0),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    return [t, Spacer(1, 0.8*mm)]

def render_checklist(text, part):
    pc = PART_COLORS.get(part, C["dark_green"])
    cb_cell = Table([
        [Paragraph("☐", ps("_cb", size=10, leading=13, color=pc, bold=True, align=TA_CENTER))],
        [Paragraph("CHECK-<br/>LIST", ps("_cblbl", size=5, leading=6.5,
                    color=pc, bold=True, align=TA_CENTER))],
    ], colWidths=[8*mm])
    cb_cell.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(0,0), 4),("BOTTOMPADDING",(0,0),(0,0), 0),
        ("TOPPADDING",(0,1),(0,1), 0),("BOTTOMPADDING",(0,1),(0,1), 4),
        ("LEFTPADDING",(0,0),(-1,-1), 0),("RIGHTPADDING",(0,0),(-1,-1), 0),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))
    data = [[cb_cell, Paragraph(text, S["check"])]]
    t = Table(data, colWidths=[8*mm, FRAME_W - 8*mm])
    t.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(0,0),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("RIGHTPADDING",(1,0),(1,0),4),("LEFTPADDING",(1,0),(1,0),6),
        ("LINEBELOW",(0,0),(-1,-1),0.3,C["light_grey"]),
    ]))
    return [t]

def render_template(text):
    data = [
        [Paragraph("TEMPLATE", ps("_tl",size=7.5,leading=9,color=C["white"],bold=True))],
        [Paragraph(text.replace("\n","<br/>"), S["tmpl"])],
    ]
    t = Table(data, colWidths=[FRAME_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0), C["purple"]),
        ("BACKGROUND",(0,1),(0,1), C["light_purple"]),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LINEBELOW",(0,0),(0,0),1,C["purple"]),
        ("BOX",(0,0),(-1,-1),0.5,C["purple"]),
    ]))
    return [t, Spacer(1,2*mm)]

def render_case(text):
    """Case study teaser — short text only. Full case study accessed via PEER-CONNECT block below."""
    data = [
        [Paragraph("📖  CASE STUDY", ps("_cl",size=7.5,leading=9,color=C["white"],bold=True))],
        [Paragraph(text, S["case"])],
        [Paragraph("Want to learn more or speak to the platform involved? See the peer connection link below.",
                   ps("_cfoot", size=7.5, leading=10, color=C["mid_grey"], italic=True))],
    ]
    t = Table(data, colWidths=[FRAME_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0), C["brown"]),
        ("BACKGROUND",(0,1),(0,2), C["light_brown"]),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("BOX",(0,0),(-1,-1),0.5,C["brown"]),
    ]))
    return [t, Spacer(1,1*mm)]

def _linkify_refs(text):
    """Convert cross-references to internal Annex/Part/Template destinations into
    clickable PDF links.  Must be applied AFTER XML-escaping (the function
    applies _xesc internally, then adds <link> markup on top)."""
    t = _xesc(text)
    # Ordered so longer matches win (Annex A before bare 'A', etc.)
    _LINK = [
        # Annexes ── anchor IDs mirror anchor_id(section_text) below
        (r"Annex A\b",
         '<link dest="s_Annex_A__Legal_Pro_Bono_Support"'
         ' color="#00424D"><u>Annex A</u></link>'),
        (r"Annex B\b",
         '<link dest="s_Annex_B__Emergency_Grants_Mechanisms"'
         ' color="#00424D"><u>Annex B</u></link>'),
        (r"Annex C\b",
         '<link dest="s_Annex_C__Physical___Digital_Security_Support"'
         ' color="#00424D"><u>Annex C</u></link>'),
        # Named sections
        (r"Part 3 \(Legal Support\)",
         '<link dest="s_3__Legal_Support" color="#00424D">'
         '<u>Part 3 (Legal Support)</u></link>'),
        (r"Part 4 \(Emergency Funding\)",
         '<link dest="s_4__Emergency_Funding" color="#00424D">'
         '<u>Part 4 (Emergency Funding)</u></link>'),
        (r"Part 5 \(Safe Comms\)",
         '<link dest="s_5__Safe_Comms" color="#00424D">'
         '<u>Part 5 (Safe Comms)</u></link>'),
        # Solidarity request template
        (r"solidarity request template",
         '<link dest="tmpl_P2_TEMPLATE_001" color="#00424D">'
         '<u>solidarity request template</u></link>'),
    ]
    for pattern, repl in _LINK:
        t = re.sub(pattern, repl, t, flags=re.IGNORECASE)
    return t


def render_db_ref(text, last_verified):
    lv_str = f"Last verified: {last_verified}" if last_verified else ""
    data = [[
        Paragraph("↗  LIVE DATABASE REFERENCE", ps("_drl",size=7,leading=9,
                   color=C["blue"],bold=True)),
        Paragraph(lv_str, ps("_drv",size=7,leading=9,color=C["mid_grey"],align=TA_RIGHT)),
    ],[
        Paragraph(_linkify_refs(text), S["dbref"]), "",
    ]]
    t = Table(data, colWidths=[FRAME_W*0.65, FRAME_W*0.35])
    t.setStyle(TableStyle([
        ("SPAN",(0,1),(1,1)),
        ("BACKGROUND",(0,0),(-1,-1), C["light_blue"]),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("BOX",(0,0),(-1,-1),0.75,C["blue"]),
        ("LINEBELOW",(0,0),(-1,0),0.5,C["blue"]),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    return [t, Spacer(1,2*mm)]

def _xesc(text):
    """Escape XML/HTML special characters so ReportLab Paragraph doesn't crash."""
    return (str(text or "")
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;"))


def render_mechanism_card(mech):
    if not mech:
        return [Paragraph("[Mechanism not found]", S["over"]), Spacer(1,2*mm)]

    status      = str(mech.get("status", "VERIFY"))
    pe          = str(mech.get("platform_eligible", "PARTIAL"))
    org         = _xesc(mech.get("organisation", ""))
    name        = _xesc(mech.get("mechanism_name", ""))
    lv          = _xesc(mech.get("last_verified", ""))
    url         = str(mech.get("db_url", "") or "")   # used raw in href; display copy escaped
    url_display = _xesc(url)
    elig        = _xesc(mech.get("eligibility_note", ""))
    access      = _xesc(mech.get("how_to_access", ""))
    timeframe   = _xesc(mech.get("timeframe", ""))
    constraints = _xesc(mech.get("constraints", ""))

    status_colors = {
        "ACTIVE":     (C["mid_green"],  C["light_green"]),
        "RESTRICTED": (C["orange"],     C["light_amber"]),
        "CLOSED":     (C["deep_red"],   C["light_red"]),
        "VERIFY":     (C["amber"],      h("FFF0CC")),
    }
    sc, sc_bg = status_colors.get(status, (C["grey"], C["light_grey"]))
    pe_c = C["mid_green"] if pe=="YES" else C["orange"] if pe=="PARTIAL" else C["deep_red"]

    w3 = FRAME_W / 3

    # Header row: name + badges
    badge_row = [[
        Paragraph(name, S["mname"]),
        Table([[Paragraph(f"PLATFORMS: {pe}",
                          ps("_pe",size=7,leading=9,color=C["white"],bold=True,align=TA_CENTER))]],
              colWidths=[26*mm], rowHeights=[6*mm],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1),pe_c),
                                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                *[("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]])),
        Table([[Paragraph(status,
                          ps("_st",size=7,leading=9,color=C["white"],bold=True,align=TA_CENTER))]],
              colWidths=[18*mm], rowHeights=[6*mm],
              style=TableStyle([("BACKGROUND",(0,0),(-1,-1),sc),
                                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                *[("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3)]])),
    ]]
    hdr_t = Table(badge_row, colWidths=[FRAME_W-48*mm, 26*mm, 18*mm])
    hdr_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),C["grey"]),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))

    # Three-column detail
    def field_col(label, body):
        return [Paragraph(label, S["mlabel"]), Paragraph(body, S["mfield"])]
    
    detail_data = [[field_col("HOW TO ACCESS", access),
                    field_col("TIMEFRAME", timeframe),
                    field_col("CONSTRAINTS", constraints)]]
    detail_t = Table(detail_data, colWidths=[w3, w3, w3])
    detail_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), C["light_grey"]),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LINEBEFORE",(1,0),(1,0),0.5,C["mid_grey"]),
        ("LINEBEFORE",(2,0),(2,0),0.5,C["mid_grey"]),
    ]))

    # Eligibility note
    elig_t = Table([[Paragraph("PLATFORM ELIGIBILITY NOTE", S["mlabel"]),
                     Paragraph(elig, S["mfield"])]],
                   colWidths=[40*mm, FRAME_W-40*mm])
    elig_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), sc_bg),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))

    # Footer
    footer_t = Table([[Paragraph(
        f"{org}  ·  Last verified: {lv}  ·  <a href='{url}' color='blue'>{url_display}</a>",
        ps("_mf",size=7,leading=9,color=C["mid_grey"]))]], colWidths=[FRAME_W])
    footer_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),C["white"]),
        ("LEFTPADDING",(0,0),(-1,-1),5),("TOPPADDING",(0,0),(-1,-1),3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LINEABOVE",(0,0),(-1,-1),0.5,C["mid_grey"]),
    ]))

    outer = Table([[hdr_t],[detail_t],[elig_t],[footer_t]], colWidths=[FRAME_W])
    outer.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.75,C["grey"]),
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    return [outer, Spacer(1,3*mm)]

def render_region_nav(text, section):
    """Pipe-separated region names → row of clickable buttons linking to annex anchors.
    content_text format: 'Africa | Asia-Pacific | Europe | Latin America & Caribbean | Pacific'
    """
    regions = [r.strip() for r in text.split("|") if r.strip()]
    if not regions:
        return []

    def region_slug(r):
        return "reg_" + re.sub(r'[^a-zA-Z0-9]', '_', r.lower()).strip("_")

    btn_w = FRAME_W / max(len(regions), 1)
    cells = []
    for r in regions:
        aid = region_slug(r)
        cells.append(Paragraph(
            f'<link dest="{aid}" color="#FFFFFF">{r}</link>',
            ps("_rnav", size=8, leading=11, color=C["white"], bold=True, align=TA_CENTER)))

    t = Table([cells], colWidths=[btn_w]*len(regions), rowHeights=[9*mm])
    cmds = [
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("LINEBEFORE",(1,0),(-1,-1),0.5,C["white"]),
    ]
    for i in range(len(regions)):
        bg = C["forus_dark"] if i % 2 == 0 else C["forus_mint"]
        cmds.append(("BACKGROUND",(i,0),(i,0), bg))
    t.setStyle(TableStyle(cmds))

    label = Table([[Paragraph("JUMP TO REGION IN ANNEX  ↓",
                               ps("_rnlbl", size=6.5, leading=9,
                                  color=C["mid_grey"], bold=True))]],
                  colWidths=[FRAME_W])
    label.setStyle(TableStyle([
        ("LEFTPADDING",(0,0),(-1,-1),0),("TOPPADDING",(0,0),(-1,-1),3),
        ("BOTTOMPADDING",(0,0),(-1,-1),2),
    ]))
    return [Spacer(1,3*mm), label, t, Spacer(1,4*mm)]


def render_country_entry(text, part):
    """Annex country entry.  content_text format: 'Country Name | Provider details…'
    The country name becomes a named PDF anchor and a bold header.
    """
    pc = PART_COLORS.get(part, C["dark_green"])
    if "|" in text:
        country, details = text.split("|", 1)
        country = country.strip()
        details = details.strip()
    else:
        country = text.strip()
        details = ""

    caid = "ctry_" + re.sub(r'[^a-zA-Z0-9]', '_', country.lower()).strip("_")

    hdr = Table([[
        SectionAnchor(caid),   # invisible anchor at country level
        Paragraph(country.upper(), ps("_ch", size=8.5, leading=11,
                                       color=C["white"], bold=True)),
    ]], colWidths=[0, FRAME_W])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), pc),
        ("LEFTPADDING",(0,0),(0,0),0),("RIGHTPADDING",(0,0),(0,0),0),
        ("LEFTPADDING",(1,0),(1,0),8),("RIGHTPADDING",(1,0),(1,0),6),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))

    body_rows = [[hdr]]
    if details:
        body_rows.append([Paragraph(details, ps("_cd", size=8.5, leading=13, color=C["grey"]))])

    outer = Table(body_rows, colWidths=[FRAME_W])
    outer.setStyle(TableStyle([
        ("BACKGROUND",(0,1),(-1,-1), C["light_grey"]) if details else ("BACKGROUND",(0,0),(-1,-1), C["light_grey"]),
        ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("LEFTPADDING",(0,1),(0,1),8) if details else ("TOPPADDING",(0,0),(0,0),0),
        ("RIGHTPADDING",(0,1),(0,1),8) if details else ("TOPPADDING",(0,0),(0,0),0),
        ("TOPPADDING",(0,1),(0,1),5) if details else ("TOPPADDING",(0,0),(0,0),0),
        ("BOTTOMPADDING",(0,1),(0,1),6) if details else ("TOPPADDING",(0,0),(0,0),0),
        ("BOX",(0,0),(-1,-1),0.5, pc),
    ]))
    return [outer, Spacer(1,2*mm)]


def render_peer_connect(text, ref=None):
    """Peer insight card.  content_text is an organisation description (v2.2+).
    Personal names are stripped so only the organisation/context is shown.
    Legacy URLs are also accepted and displayed as a link instead.
    """
    import re as _re
    text = (text or "").strip()
    # Detect legacy URL content
    is_url = text.startswith("http") or text.startswith("www.")
    if is_url:
        url = text or "https://forus-international.org/peer-connect"
        if ref:
            sep = "&" if "?" in url else "?"
            url = f"{url}{sep}ref={ref}"
        body_para = Paragraph(
            f'<a href="{url}" color="#D4F0F1"><u>{url}</u></a>',
            ps("_pcurl", size=8, leading=11, color=C["forus_teal_lt"]))
    else:
        # Strip leading personal name: one or more Title-case words followed by ", "
        # e.g. "Moses Isooba, " → "" or "Shannon Kindornay, " → ""
        clean = _re.sub(r'^(?:[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*),\s*', '', text)
        body_para = Paragraph(clean, ps("_pcbody", size=8.5, leading=12,
                                         color=C["forus_teal_lt"]))

    data = [[
        Paragraph("🤝", ps("_pci", size=12, leading=14, color=C["white"],
                             align=TA_CENTER)),
        Table([
            [Paragraph("PEER INSIGHT",
                        ps("_pclbl", size=7.5, leading=10, color=C["white"], bold=True))],
            [body_para],
        ], colWidths=[FRAME_W - 16*mm]),
    ]]
    t = Table(data, colWidths=[16*mm, FRAME_W - 16*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C["forus_dark"]),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (0, 0), 4), ("RIGHTPADDING", (0, 0), (0, 0), 4),
        ("LEFTPADDING",  (1, 0), (1, 0), 0), ("RIGHTPADDING", (1, 0), (1, 0), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
        ("LINEAFTER",    (0, 0), (0, 0),  1, C["forus_teal"]),
    ]))
    return [Spacer(1, 3*mm), t, Spacer(1, 3*mm)]


# ── Page chrome ───────────────────────────────────────────────────────────────

def anchor_id(section_text):
    """Create a safe, unique PDF anchor name from section text."""
    return "s_" + re.sub(r'[^a-zA-Z0-9]', '_', str(section_text))[:48]


class SectionAnchor(Flowable):
    """Zero-height flowable that registers a named PDF destination.
    Placed at each section so ToC links can jump directly to it."""
    def __init__(self, name):
        super().__init__()
        self.name   = name
        self.width  = 0
        self.height = 0
    def draw(self):
        # bookmarkHorizontal sets the destination at the current canvas position
        self.canv.bookmarkHorizontal(self.name, 0, 0)


class SetMeta(Flowable):
    """Zero-size flowable placed immediately AFTER each section banner.
    Pass 1 of the two-pass build records which page each one lands on.
    Pass 2 uses that map to draw chrome with the correct section label."""
    def __init__(self, part, section):
        super().__init__()
        self.part    = part
        self.section = section
        self.width   = 0
        self.height  = 0
    def draw(self):
        pass


class ToolkitDoc(BaseDocTemplate):
    def __init__(self, fn, access_level=1, page_map=None, **kw):
        super().__init__(fn, **kw)
        self.access_level = access_level
        # page_map: {page_number: (part, section)} supplied from pass 1.
        # None means this is pass 1 — collect rather than render.
        self._page_map  = page_map or {}
        self._meta_log  = []   # populated during pass 1 via afterFlowable
        f = Frame(FRAME_X, MB, FRAME_W, PAGE_H - MT - MB, id="main",
                  showBoundary=0,
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
        self.addPageTemplates([PageTemplate(id="main", frames=[f],
                                            onPage=self._chrome)])

    def afterFlowable(self, flowable):
        """Called after each flowable is placed. Used in pass 1 only."""
        if isinstance(flowable, SetMeta):
            self._meta_log.append((self.page, flowable.part, flowable.section))

    def _lookup(self, page_num):
        """Return (part, section) for page_num using the pass-1 page map."""
        result = (1, "")
        for p in sorted(self._page_map):
            if p <= page_num:
                result = self._page_map[p]
        return result

    def _chrome(self, canv, doc):
        canv.saveState()
        pn = canv.getPageNumber()
        part, section = doc._lookup(pn)
        pc         = PART_COLORS.get(part, C["dark_green"])
        part_label = PART_LABELS.get(part, "")

        # ── Left navigation strip ────────────────────────────────────────────
        canv.setFillColor(pc)
        canv.roundRect(ML, MB, NAV_W, PAGE_H - MT - MB, 3, fill=1, stroke=0)
        canv.saveState()
        canv.translate(ML + NAV_W / 2, PAGE_H / 2)
        canv.rotate(90)
        canv.setFillColor(C["white"])
        canv.setFont("Helvetica-Bold", 6.5)
        canv.drawCentredString(0, -2.5, part_label)
        canv.restoreState()

        # ── Running header ───────────────────────────────────────────────────
        canv.setStrokeColor(C["mid_grey"])
        canv.setLineWidth(0.4)
        canv.line(FRAME_X, PAGE_H - MT + 2*mm, PAGE_W - MR, PAGE_H - MT + 2*mm)
        canv.setFillColor(C["mid_grey"])
        canv.setFont("Helvetica", 7)
        canv.drawString(FRAME_X, PAGE_H - MT + 3.5*mm,
                        "Forus Resilience & Support Toolkit")
        canv.setFont("Helvetica-Bold", 7)
        canv.drawRightString(PAGE_W - MR, PAGE_H - MT + 3.5*mm, section)

        # ── Footer ───────────────────────────────────────────────────────────
        canv.setStrokeColor(C["mid_grey"])
        canv.line(FRAME_X, MB - 4*mm, PAGE_W - MR, MB - 4*mm)
        canv.setFillColor(C["mid_grey"])
        canv.setFont("Helvetica", 7)
        al = "PUBLIC VERSION" if doc.access_level == 1 else "FORUS NETWORK — CONFIDENTIAL"
        canv.drawString(FRAME_X, MB - 7*mm, al)
        canv.drawCentredString(PAGE_W / 2, MB - 7*mm,
                               f"v{VERSION}  ·  {DATE_STAMP}")
        canv.drawRightString(PAGE_W - MR, MB - 7*mm, f"p. {pn}")
        canv.restoreState()


# ── Load spreadsheet ──────────────────────────────────────────────────────────
def load_data(access_level, language="EN"):
    """Load CONTENT rows filtered by language and access level, plus ANNEXES.

    language: "EN" (default), "FR", or "ES".
    For FR/ES, ANNEX translatable fields are swapped for their _fr / _es column values.
    """
    lang = language.upper() if language else "EN"
    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)

    ws_c = wb["CONTENT"]
    hdrs = [c.value for c in ws_c[2]]
    cm   = {h:i for i,h in enumerate(hdrs) if h}
    rows = []
    for row in ws_c.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        item = {h: row[i] for h,i in cm.items() if i<len(row)}
        if not item.get("block_id"): continue
        # Filter by language
        row_lang = str(item.get("language", "EN") or "EN").upper()
        if row_lang != lang:
            continue
        if int(item.get("sensitivity",1) or 1) <= access_level:
            rows.append(item)

    ws_m = wb["ANNEXES"]
    mhdrs = [c.value for c in ws_m[2]]
    mcm   = {h:i for i,h in enumerate(mhdrs) if h}
    mechs = {}
    # Fields that have translated variants in the spreadsheet
    _translatable = ("eligibility_note", "how_to_access", "timeframe", "constraints", "notes")
    lang_suffix = f"_{lang.lower()}" if lang != "EN" else None
    for row in ws_m.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        m = {h: row[i] for h,i in mcm.items() if i<len(row)}
        if m.get("mech_id"):
            # Swap in translated field values when language != EN
            if lang_suffix:
                for field in _translatable:
                    translated = m.get(f"{field}{lang_suffix}")
                    if translated:
                        m[field] = translated
            mechs[m["mech_id"]] = m

    def sort_key(x):
        part_val = x.get("part", 1)
        part_int = int(part_val) if part_val is not None else 1

        do_val = x.get("display_order")
        # Guard against falsy-zero bug: 0 is a valid order, must not fall back to 99
        if do_val is None or do_val == "":
            display_ord = 99
        else:
            display_ord = int(do_val)

        # Sort purely by display_order — no global SUBSECTION float.
        # Each SUBSECTION's display_order is set below the entries that follow it,
        # so regions stay grouped with their country entries (e.g. Africa SUBSECTION
        # at order 0, Africa entries at 1-3, Asia SUBSECTION at 20, entries at 21-23…)
        return (
            part_int,
            str(x.get("section", "")),
            TIME_ORDER.get(str(x.get("time_horizon", "general") or "general"), 5),
            display_ord,
        )

    rows.sort(key=sort_key)
    return rows, mechs


# ── Tools data (TOOLS sheet) ──────────────────────────────────────────────────

def load_tools_data():
    """Load all TOOLS sheet entries into a flat dict {field_key: content_text}.
    Returns an empty dict if the sheet does not exist yet.
    """
    try:
        wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    except Exception:
        return {}
    if "TOOLS" not in wb.sheetnames:
        return {}
    ws = wb["TOOLS"]
    hdrs = [c.value for c in ws[2]]
    try:
        ki = hdrs.index("field_key")
        vi = hdrs.index("content_text")
    except ValueError:
        return {}
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[ki]: continue
        data[str(row[ki]).strip()] = str(row[vi] or "").strip() if row[vi] is not None else ""
    return data


def ensure_tools_sheet(wb):
    """Add a TOOLS sheet to the workbook with all default content if it doesn't exist.
    No-ops if the sheet already exists. Returns the worksheet."""
    from forus_tools_v4 import T1_DEFAULTS, T2_DEFAULTS, T3_DEFAULTS, T4_DEFAULTS, TOOL_LABELS
    from forus_appendix_tools import A1_DEFAULTS, A2_DEFAULTS, A3_DEFAULTS, APPENDIX_LABELS

    if "TOOLS" in wb.sheetnames:
        return wb["TOOLS"]

    ws = wb.create_sheet("TOOLS")
    # Header rows
    ws.cell(1, 1, "FORUS TOOLKIT — TOOLS CONTENT")
    ws.cell(2, 1, "tool_id");    ws.cell(2, 2, "tool_name")
    ws.cell(2, 3, "field_key");  ws.cell(2, 4, "field_label")
    ws.cell(2, 5, "content_text"); ws.cell(2, 6, "word_limit")
    ws.cell(2, 7, "notes"); ws.cell(2, 8, "last_updated")
    ws.cell(2, 9, "change_flag")

    # Word limits per field type (rough guide)
    _WL = {
        "WHY_THIS_MATTERS": 50, "HOW_TO_USE": 40, "IF_YOU_FIND_GAPS": 40,
        "IF_ANY_NO": 40, "BUYING_TIME": 30, "PROACTIVE_LEGAL_HEALTH": 50,
        "CHECKLIST": 20, "SCORE": 20, "OUTCOME": 40, "BOX_BODY": 55,
        "TIER_DESC": 50, "TIER_IND": 20, "GATE_ITEM": 20,
        "POINT_DESC": 25, "NOTE": 40, "SUB": 20, "STAY_QUIET": 35,
        "PAUSE": 35, "GO_PUBLIC": 35, "USE_INSTRUCTION": 30,
        "ITEM": 25, "ITEM_NOTE": 25,
    }

    def _wl(key):
        for k, v in _WL.items():
            if k in key.upper(): return v
        return 30

    today = datetime.date.today().isoformat()
    tool_groups = [
        ("T1", TOOL_LABELS["T1"], T1_DEFAULTS),
        ("T2", TOOL_LABELS["T2"], T2_DEFAULTS),
        ("T3", TOOL_LABELS["T3"], T3_DEFAULTS),
        ("T4", TOOL_LABELS["T4"], T4_DEFAULTS),
        ("A1", APPENDIX_LABELS["A1"], A1_DEFAULTS),
        ("A2", APPENDIX_LABELS["A2"], A2_DEFAULTS),
        ("A3", APPENDIX_LABELS["A3"], A3_DEFAULTS),
    ]
    r = 3
    for tid, tname, defaults in tool_groups:
        for fkey, fval in defaults.items():
            ws.cell(r, 1, tid); ws.cell(r, 2, tname)
            ws.cell(r, 3, fkey)
            # Make a readable label from the key
            label = fkey.replace(f"{tid}_", "").replace("_", " ").title()
            ws.cell(r, 4, label)
            ws.cell(r, 5, fval)
            ws.cell(r, 6, _wl(fkey))
            ws.cell(r, 7, "AI-updatable field — keep concise, check context carefully")
            ws.cell(r, 8, today)
            ws.cell(r, 9, "OK")
            r += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14

    return ws


# ── PDF merging (tool pages appended after main toolkit) ─────────────────────

# Which part each tool belongs to — tool pages are inserted after that part's last page
_TOOL_PART = {
    "T1": 1,   # Compliance Self-Check    → Part 1  (B1 Crisis Scenarios)
    "A1": 2,   # Platform Role Clarifier  → Part 2  (B2 Solidarity)
    "T2": 3,   # Legal Decision Tree      → Part 3  (B3 Legal Support)
    "A3": 4,   # Emergency Funding Nav.   → Part 4  (B4 Emergency Funding)
    "T3": 5,   # Go public / stay quiet   → Part 5  (B5 Safe Comms)
    "T4": 5,   # Do-No-Harm Checklist     → Part 5  (B5 Safe Comms)
    "A2": 6,   # Diversification Gate     → Part 6  (B6 Diversification)
}


def _build_tool_buf_for_ids(ids, tools_data):
    """Return a BytesIO with pages for the given tool/appendix IDs, in canonical order."""
    t_ids = [t for t in ["T1", "T2", "T3", "T4"] if t in ids]
    a_ids = [t for t in ["A1", "A2", "A3"]         if t in ids]
    bufs = []
    if t_ids:
        try:
            from forus_tools_v4 import build_tools_pdf
            b = build_tools_pdf(t_ids, data=tools_data)
            if b: bufs.append(b)
        except Exception as e:
            print(f"  ⚠ Could not build tool pages {t_ids}: {e}")
    if a_ids:
        try:
            from forus_appendix_tools import build_appendix_pdf
            b = build_appendix_pdf(a_ids, data=tools_data)
            if b: bufs.append(b)
        except Exception as e:
            print(f"  ⚠ Could not build appendix pages {a_ids}: {e}")
    return bufs


def _merge_tools_inline(main_path, tools_selection, page_map, out_path):
    """Insert tool pages inline — each tool appears right after its section's last page.

    page_map: {page_number (1-indexed): (part_int, section_str)}  from pass-1 build.
    If pypdf is unavailable the main PDF is simply renamed/copied to out_path.
    """
    if not tools_selection:
        if main_path != out_path:
            import shutil; shutil.move(main_path, out_path)
        return

    selected = [tid for tid in _TOOL_PART if tools_selection.get(tid)]
    if not selected:
        if main_path != out_path:
            import shutil; shutil.move(main_path, out_path)
        return

    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfWriter, PdfReader
        except ImportError:
            print("  ⚠ pypdf not installed — tool pages skipped. Add pypdf to requirements.txt")
            if main_path != out_path:
                import shutil; shutil.move(main_path, out_path)
            return

    # Reverse page_map → part → last page number
    part_last_page = {}
    for pgnum, (part, _) in page_map.items():
        part_last_page[part] = max(part_last_page.get(part, 0), pgnum)

    # Group selected tools by part, preserving canonical draw order within each part
    tools_data = load_tools_data()
    part_tool_ids = {}
    for tid in selected:
        p = _TOOL_PART[tid]
        part_tool_ids.setdefault(p, []).append(tid)

    # Pre-build one BytesIO per part group
    part_bufs = {}
    for p, ids in part_tool_ids.items():
        raw_bufs = _build_tool_buf_for_ids(ids, tools_data)
        if raw_bufs:
            # Combine multiple source buffers into one merged buf for this part
            w = PdfWriter()
            for rb in raw_bufs:
                rb.seek(0)
                for pg in PdfReader(rb).pages:
                    w.add_page(pg)
            combined = io.BytesIO()
            w.write(combined)
            combined.seek(0)
            part_bufs[p] = combined

    if not part_bufs:
        if main_path != out_path:
            import shutil; shutil.move(main_path, out_path)
        return

    # Walk through main PDF; after each page check if tool pages should follow
    writer = PdfWriter()
    with open(main_path, "rb") as f:
        reader = PdfReader(f)
        for i, page in enumerate(reader.pages):
            pg_num = i + 1  # 1-indexed
            writer.add_page(page)
            # Insert tool pages for any part whose last page is this page
            # Sort by part to keep tools in section order
            for part in sorted(part_bufs.keys()):
                if part_last_page.get(part, 0) == pg_num:
                    buf = part_bufs[part]
                    buf.seek(0)
                    for tpg in PdfReader(buf).pages:
                        writer.add_page(tpg)

    with open(out_path, "wb") as f:
        writer.write(f)
    if main_path != out_path:
        try: os.remove(main_path)
        except OSError: pass

# ── Cover & ToC ───────────────────────────────────────────────────────────────

# Part intro lines shown in the ToC (one per part)
PART_INTROS = {
    1: "Crisis response guides — legislative, funding, digital & reputational",
    2: "Activating peer solidarity across the Forus network",
    3: "Finding and accessing legal support when it matters",
    4: "Emergency funding mechanisms for platforms under pressure",
    5: "Safe advocacy, communications & digital security",
    6: "Diversification, mutualisation & long-term sustainability",
    7: "Keeping this toolkit accurate and up to date",
    8: "Regional and country-level resource directories — legal, funding & digital support",
}

def build_cover(story, access_level, sections_in_order=None, page_map=None):
    """Build the cover page.
    sections_in_order: list of (part, section_text) in reading order — used for ToC.
    page_map: {page: (part, section)} from pass 1 — None during pass 1 itself.
    """
    label = "PUBLIC VERSION" if access_level==1 else "FORUS NETWORK — CONFIDENTIAL"
    cover_rows = [
        [Paragraph("FORUS RESILIENCE &amp;<br/>SUPPORT TOOLKIT", S["cover_t"])],
        [Paragraph("Navigating Legal, Solidarity Support &amp; Sustainable Resource Models", S["cover_s"])],
        [Spacer(1,6*mm)],
        [Paragraph(label, S["cover_m"])],
        [Paragraph(f"Updated {DATE_STAMP}  ·  Flourish Nonprofits for Forus International", S["cover_m"])],
    ]
    ct = Table(cover_rows, colWidths=[FRAME_W])
    ct.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), C["dark_green"]),
        ("TOPPADDING",(0,0),(0,0), 35*mm),
        ("TOPPADDING",(0,1),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))
    story.append(ct)
    story.append(PageBreak())

    # ── Table of Contents page ────────────────────────────────────────────────
    # Build a reverse map: section_text → page number (from pass 1 page_map)
    section_to_page = {}
    if page_map:
        for pg, (pt, sec) in sorted(page_map.items()):
            if sec and sec not in section_to_page:
                section_to_page[sec] = pg

    story.append(Paragraph("CONTENTS", ps("_toc_title", size=14, leading=18,
                            color=C["dark_green"], bold=True)))
    story.append(Spacer(1, 4*mm))

    # Part colours for the dot/rule beside each part heading
    current_part = None
    toc_rows = []

    if sections_in_order:
        for part, section in sections_in_order:
            pc = PART_COLORS.get(part, C["dark_green"])

            # Part heading row — shown once per part
            if part != current_part:
                current_part = part
                pl = PART_LABELS.get(part, f"PART {part}")
                intro = PART_INTROS.get(part, "")

                # Part label cell
                part_cell = Table(
                    [[Paragraph(f"PART {part}", ps("_tp_lbl", size=7, leading=9,
                                color=C["white"], bold=True, align=TA_CENTER))]],
                    colWidths=[14*mm], rowHeights=[5.5*mm])
                part_cell.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,-1), pc),
                    ("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1),
                    ("LEFTPADDING",(0,0),(-1,-1),2),("RIGHTPADDING",(0,0),(-1,-1),2),
                ]))

                toc_rows.append([
                    part_cell,
                    Paragraph(pl, ps("_tp_name", size=10, leading=13,
                                     color=pc, bold=True)),
                    Paragraph(intro, ps("_tp_intro", size=8, leading=11,
                                        color=C["mid_grey"], italic=True)),
                    Paragraph(""),
                ])

            # Section row — with clickable link in pass 2
            aid = anchor_id(section)
            pg_num = section_to_page.get(section)

            if pg_num:
                # Pass 2: real page number + clickable link
                link_style = ps("_toc_link", size=9, leading=13,
                                color=C["forus_dark"])
                sec_para = Paragraph(
                    f'<link dest="{aid}" color="#00424D">{section}</link>',
                    link_style)
                pg_para  = Paragraph(
                    f'<link dest="{aid}" color="#00424D">{pg_num}</link>',
                    ps("_toc_pg", size=9, leading=13,
                       color=C["forus_dark"], bold=True, align=TA_RIGHT))
            else:
                # Pass 1: no page numbers yet — render plain text
                sec_para = Paragraph(section,
                                     ps("_toc_s", size=9, leading=13, color=C["grey"]))
                pg_para  = Paragraph("—",
                                     ps("_toc_p", size=9, leading=13,
                                        color=C["mid_grey"], align=TA_RIGHT))

            toc_rows.append([
                Paragraph(""),   # indent cell (no part badge for sections)
                sec_para,
                Paragraph(""),   # no intro for individual sections
                pg_para,
            ])

    if toc_rows:
        col_w = [16*mm, FRAME_W*0.52, FRAME_W*0.32, 14*mm]
        toc_t = Table(toc_rows, colWidths=col_w)
        cmds = [
            ("LEFTPADDING",(0,0),(-1,-1),3),
            ("RIGHTPADDING",(0,0),(-1,-1),3),
            ("TOPPADDING",(0,0),(-1,-1),3),
            ("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]
        # Stripe section rows; part rows get a light top border
        for i, (part, section) in enumerate(sections_in_order or []):
            # Offset by 1 for each part heading that precedes this section
            pass
        # Simpler: just zebra all rows
        for i in range(len(toc_rows)):
            bg = C["light_grey"] if i % 2 == 0 else C["white"]
            cmds.append(("BACKGROUND",(0,i),(-1,i), bg))
        toc_t.setStyle(TableStyle(cmds))
        story.append(toc_t)
    else:
        story.append(Paragraph("Contents will appear here after the first build.",
                                ps("_toc_ph", size=9, color=C["mid_grey"], italic=True)))

    story.append(PageBreak())

    # ── Acronyms / Abbreviations page ────────────────────────────────────────
    story += render_acronyms_page()


# ── Acronyms page ─────────────────────────────────────────────────────────────
def render_acronyms_page():
    """One-page glossary of abbreviations used in the toolkit."""
    ACRONYMS = [
        ("ANC Peru",  "Asociación Nacional de Centros de Investigación, Promoción Social y Desarrollo"),
        ("BRT",       "Building Responses Together (CIVICUS partner network)"),
        ("CIVICUS",   "World Alliance for Citizen Participation"),
        ("CNONGD",    "Conseil National des ONG de Développement (DRC)"),
        ("CONCORD",   "European NGO Confederation for Relief and Development"),
        ("CSO",       "Civil Society Organisation"),
        ("DDP",       "Democratic Dialogue Programme (European funding)"),
        ("EU SEE",    "EU Civil Society Support Programme for South-East Europe"),
        ("FALE",      "Facility Aiding Locally Led Engagement (PIANGO model)"),
        ("FATF",      "Financial Action Task Force (global anti-money-laundering body)"),
        ("FLD",       "Front Line Defenders"),
        ("ICNL",      "International Center for Not-for-Profit Law"),
        ("LAPAS",     "Latvijas Pilsoniskā alianse — Latvian Civic Alliance"),
        ("MFF",       "EU Multi-Annual Financial Framework"),
        ("NGO",       "Non-Governmental Organisation"),
        ("NNNGO",     "Network of Networks for NGOs, Nigeria"),
        ("ODA",       "Official Development Assistance"),
        ("PDA",       "Pakistan Development Alliance"),
        ("PIANGO",    "Pacific Islands Association of Non-Governmental Organisations"),
        ("PILnet",    "Public Interest Law Network"),
        ("SLAPP",     "Strategic Lawsuit Against Public Participation"),
        ("UPR",       "Universal Periodic Review (UN Human Rights Council mechanism)"),
        ("USAID",     "United States Agency for International Development"),
    ]

    story = []
    story.append(Paragraph("ABBREVIATIONS &amp; ACRONYMS", ps(
        "_acr_title", size=12, leading=16, color=C["dark_green"], bold=True)))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "Abbreviations used throughout this toolkit are listed below.",
        ps("_acr_intro", size=8.5, leading=12, color=C["grey"], italic=True)))
    story.append(Spacer(1, 4*mm))

    rows = []
    for abbr, definition in ACRONYMS:
        rows.append([
            Paragraph(f"<b>{abbr}</b>",
                      ps("_acr_a", size=8.5, leading=12, color=C["dark_green"])),
            Paragraph(definition,
                      ps("_acr_d", size=8.5, leading=12, color=C["grey"])),
        ])

    t = Table(rows, colWidths=[28*mm, FRAME_W - 28*mm])
    t.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ("LINEBELOW",    (0,0), (-1,-2), 0.25, C["light_grey"]),
    ]))
    story.append(t)
    story.append(PageBreak())
    return story


# ── Main render ───────────────────────────────────────────────────────────────
def render_block(item, mechs, story, warnings):
    btype    = str(item.get("block_type","")).strip()
    text_raw = str(item.get("content_text","") or "")
    part     = int(item.get("part",1) or 1)
    section  = str(item.get("section",""))
    time_h   = str(item.get("time_horizon","general") or "general")
    last_upd = str(item.get("last_updated","") or "")
    bid      = str(item.get("block_id",""))
    pc       = PART_COLORS.get(part, C["dark_green"])

    text, truncated = trim(text_raw, get_limit(item))
    if truncated:
        warnings.append(f"  ⚠ OVER LIMIT: {bid} [{btype}, limit={get_limit(item)}]")

    if btype == "HEADER":
        story += render_header(text)

    elif btype == "INTRO":
        story += render_intro(text)

    elif btype == "STEP":
        if time_h in TIME_LABELS and not getattr(story[-1],"_timeline_drawn",False):
            # Only draw timeline bar at first step of each horizon group
            tl = render_timeline_bar(time_h)
            for f in tl: story.append(f)
        story += render_step(text, "·", part, truncated)

    elif btype in ("TIP","WARNING"):
        story += render_callout(text, btype, truncated)

    elif btype == "DECISION-Q":
        story.append(Spacer(1,1*mm))
        story += render_decision_q(text)

    elif btype == "DECISION-A":
        story += render_decision_a(text, truncated)

    elif btype == "CHECKLIST":
        story += render_checklist(text, part)

    elif btype == "TEMPLATE":
        # Place a named anchor so other sections can link back to this template
        tmpl_anchor = "tmpl_" + re.sub(r'[^a-zA-Z0-9]', '_', bid)
        story.append(SectionAnchor(tmpl_anchor))
        story += render_template(text)

    elif btype == "CASE":
        story += render_case(text)

    elif btype in ("PEER-CONNECTION", "PEER-CONNECT"):
        story += render_peer_connect(text_raw.strip(), ref=bid)

    elif btype == "REGION-NAV":
        story += render_region_nav(text_raw.strip(), section)

    elif btype == "COUNTRY-ENTRY":
        story += render_country_entry(text_raw.strip(), part)

    elif btype == "FEEDBACK":
        story += render_feedback(text_raw.strip())

    elif btype == "DB-REF":
        story += render_db_ref(text, last_upd)

    elif btype == "MECHANISM-REF":
        mids = re.findall(r'MECH-[A-Z]-\d+', text)
        if mids:
            for mid in mids:
                story += render_mechanism_card(mechs.get(mid))
        else:
            # No database code — render as styled arrow line with Annex hyperlinks
            story.append(Paragraph(_linkify_refs(text), S["mechref"]))
            story.append(Spacer(1,2*mm))

    else:
        story.append(Paragraph(f"[{btype}] {text}", S["normal"]))
        story.append(Spacer(1,2*mm))

def _mech_matches_regions(mech, selected_regions):
    """Return True if this mechanism's geographic_coverage includes any selected region."""
    geo = str(mech.get("geographic_coverage", "") or "").lower()
    # Always include mechanisms that cover all regions
    if any(k.lower() in geo for k in _GEO_ALL_REGIONS):
        return True
    # Check each selected region's keywords
    for rkey, v in selected_regions.items():
        if not v:
            continue
        for kw in _REGION_GEO_KEYWORDS.get(rkey, []):
            if kw.lower() in geo:
                return True
    return False


def make_story(rows, mechs, access_level, page_map=None, req=None):
    """Build the complete story list. Called twice — once per pass.
    page_map is None in pass 1; supplied from pass 1 results in pass 2.
    req: optional request dict; when supplied, annex sections are generated
         from ANNEXES data after all CONTENT rows."""
    story    = []
    warnings = []

    # Collect ordered unique (part, section) pairs for the ToC,
    # including any selected annex sections.
    sections_in_order = []
    seen_sections = set()
    for item in rows:
        sec  = str(item.get("section", ""))
        part = int(item.get("part", 1) or 1)
        if sec and sec not in seen_sections:
            seen_sections.add(sec)
            sections_in_order.append((part, sec))
    if req:
        selected_annexes = {k: v for k, v in req.get("annexes", {}).items() if v}
        for ann_sec in selected_annexes:
            if ann_sec not in seen_sections:
                sections_in_order.append((8, ann_sec))

    build_cover(story, access_level,
                sections_in_order=sections_in_order,
                page_map=page_map)

    prev_section  = None
    step_counters = {}
    # Track the active phase from HEADER blocks.
    # "BEFORE CRISIS" headers → preemptive; "DURING CRISIS" → responsive.
    # This drives the timeline bar when all rows have time_horizon="general".
    current_phase = "responsive"   # sensible default until first HEADER seen

    for item in rows:
        part    = int(item.get("part", 1) or 1)
        section = str(item.get("section", ""))
        btype   = str(item.get("block_type", "")).strip()
        time_h  = str(item.get("time_horizon", "general") or "general")

        # Section 7 (Update Guide) — network/public builds show only the
        # FEEDBACK block (a simple "tell us what's out of date" prompt).
        # Full maintenance checklists/tips are only for Secretariat internal use.
        if part == 7 and btype not in ("FEEDBACK", "PEER-CONNECT", "PEER-CONNECTION"):
            # Still render the section banner when section changes
            if section != prev_section:
                if prev_section is not None:
                    story.append(Spacer(1, 8*mm))
                for f in render_section_banner(section, part):
                    story.append(f)
                story.append(SectionAnchor(anchor_id(section)))
                story.append(SetMeta(part, section))
                prev_section = section
                prev_time    = None
                step_counters[section] = {}
                current_phase = "responsive"
                make_story._prev_display_h = None
            continue   # skip everything except FEEDBACK in Part 7

        # New section → banner, anchor for ToC links, then SetMeta for chrome
        if section != prev_section:
            if prev_section is not None:
                story.append(Spacer(1, 8*mm))
            for f in render_section_banner(section, part):
                story.append(f)
            story.append(SectionAnchor(anchor_id(section)))   # ← PDF destination
            story.append(SetMeta(part, section))               # ← chrome metadata
            prev_section  = section
            prev_time     = None
            step_counters[section] = {}
            # Reset phase context when entering a new section
            current_phase = "responsive"
            make_story._prev_display_h = None
        else:
            if "prev_time" not in dir():
                prev_time = None

        if btype == "STEP":
            key = time_h
            step_counters[section][key] = step_counters[section].get(key, 0) + 1
            item = dict(item)
            # If the row has an explicit preemptive/responsive horizon, honour it.
            # Otherwise use the phase set by the most recent HEADER block in this section.
            if time_h == "preemptive":
                display_h = "preemptive"
            elif time_h in ("responsive", "first-hour", "first-24hrs", "first-72hrs"):
                display_h = "responsive"
            else:
                # time_h is "general" (v2.2 default) — derive phase from HEADER context
                display_h = current_phase
            if display_h != getattr(make_story, "_prev_display_h", None) or time_h != prev_time:
                story.append(Spacer(1, 5*mm))
                for f in render_timeline_bar(display_h): story.append(f)
                make_story._prev_display_h = display_h
                prev_time = time_h
            text_raw = str(item.get("content_text", "") or "")
            text, truncated = trim(text_raw, get_limit(item))
            if truncated:
                warnings.append(f"  ⚠ OVER LIMIT: {item.get('block_id')} [STEP, limit={get_limit(item)}]")
            story += render_step(text, step_counters[section][key], part, truncated)

        elif btype == "SUBSECTION":
            text_raw = str(item.get("content_text", "") or "")
            text, truncated = trim(text_raw, get_limit(item))
            if truncated:
                warnings.append(f"  ⚠ OVER LIMIT: {item.get('block_id')} [SUBSECTION, limit={get_limit(item)}]")
            for f in render_subsection_divider(text, part, time_h):
                story.append(f)
            # In Part 8 (Annexes), SUBSECTION marks a region — emit reg_ anchor so REGION-NAV links work
            if part == 8:
                reg_slug = "reg_" + re.sub(r'[^a-zA-Z0-9]', '_', text.lower()).strip("_")
                story.append(SectionAnchor(reg_slug))
            step_counters[section][time_h] = 0
            prev_time = time_h

        elif btype == "HEADER":
            # Render the phase banner AND update current_phase for subsequent STEPs
            text_raw = str(item.get("content_text", "") or "")
            text, _ = trim(text_raw, get_limit(item))
            tl = text.upper()
            if "BEFORE" in tl:
                current_phase = "preemptive"
            elif "DURING" in tl:
                current_phase = "responsive"
            # Force timeline bar to redraw after the phase switch
            make_story._prev_display_h = None
            story += render_header(text)
            prev_time = None

        else:
            render_block(item, mechs, story, warnings)
            if btype not in ("DECISION-Q", "DECISION-A"):
                prev_time = None

    # ── Annex sections from ANNEXES sheet ────────────────────────────────
    if req:
        selected_annexes = req.get("annexes", {})
        selected_regions = req.get("regions", {})
        for ann_sec, ann_cat in _ANNEX_CATEGORY.items():
            if not selected_annexes.get(ann_sec):
                continue
            # Section banner (part=8 gets its own nav colour)
            story.append(Spacer(1, 8*mm))
            for f in render_section_banner(ann_sec, 8):
                story.append(f)
            story.append(SectionAnchor(anchor_id(ann_sec)))
            story.append(SetMeta(8, ann_sec))

            # Filter mechanisms by category and region
            matching = [
                m for m in mechs.values()
                if str(m.get("category", "")).strip() == ann_cat
                and _mech_matches_regions(m, selected_regions)
            ]

            if matching:
                for mech in matching:
                    story += render_mechanism_card(mech)
            else:
                story.append(Paragraph(
                    "No mechanisms matched your selected regions for this annex.",
                    ps("_no_match", size=9, leading=13, color="mid_grey", italic=True)))
                story.append(Spacer(1, 4*mm))

    return story, warnings


def update_spreadsheet(all_rows_by_id):
    """Write actual word counts back into the spreadsheet and rebuild the DASHBOARD.
    all_rows_by_id: dict of block_id → {word_count, within_limit, word_limit}
    """
    import datetime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles.differential import DifferentialStyle

    wb = openpyxl.load_workbook(SPREADSHEET)

    # Ensure TOOLS sheet exists (creates it with defaults on first run)
    ensure_tools_sheet(wb)

    ws = wb["CONTENT"]

    hdrs = [c.value for c in ws[2]]
    cm   = {h: i+1 for i, h in enumerate(hdrs) if h}  # 1-indexed for ws.cell

    from openpyxl.formatting.rule import FormulaRule

    FORM_FONT  = Font(name="Arial", color="00424D", size=9, italic=True)
    NUM_FONT   = Font(name="Arial", size=9)
    BOLD_FONT  = Font(name="Arial", size=9, bold=True)

    # Column letters for formulas
    col_F_letter = chr(64 + cm["block_type"])
    col_H_letter = chr(64 + cm["content_text"])
    col_I_letter = chr(64 + cm["word_limit"])
    col_J_letter = chr(64 + cm["word_count"])

    SWITCH_TYPES = (
        '"STEP",25,"INTRO",60,"TIP",35,"WARNING",35,'
        '"DECISION-Q",20,"DECISION-A",8,"CHECKLIST",20,'
        '"TEMPLATE",120,"CASE",80,"HEADER",10,'
        '"DB-REF",40,"MECHANISM-REF",15,"SUBSECTION",12,"FEEDBACK",120,20'
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
        bid = row[0].value
        if not bid:
            continue

        # Col I: write SWITCH formula unless there's a manual numeric override
        i_cell = ws.cell(row=row_idx, column=cm["word_limit"])
        try:
            int(i_cell.value)   # manual override — leave it alone
        except (TypeError, ValueError):
            i_cell.value = f'=SWITCH({col_F_letter}{row_idx},{SWITCH_TYPES})'
            i_cell.font  = FORM_FONT

        # Col J: live word-count formula (always refresh)
        j_cell = ws.cell(row=row_idx, column=cm["word_count"])
        j_cell.value = (
            f'=IF(TRIM({col_H_letter}{row_idx})="",0,'
            f'LEN(TRIM({col_H_letter}{row_idx}))'
            f'-LEN(SUBSTITUTE(TRIM({col_H_letter}{row_idx})," ",""))+1)'
        )
        j_cell.font = NUM_FONT

        # Col K: OVER ↑ / ✓ formula (conditional formatting handles colour)
        k_cell = ws.cell(row=row_idx, column=cm["within_limit"])
        k_cell.value = (
            f'=IF({col_J_letter}{row_idx}>{col_I_letter}{row_idx},"OVER ↑","✓")'
        )
        k_cell.font = BOLD_FONT

    # ── Re-apply conditional formatting on J and K ────────────────────────────
    last_row = ws.max_row
    range_J  = f"{col_J_letter}3:{col_J_letter}{last_row}"
    range_K  = f"{chr(64+cm['within_limit'])}3:{chr(64+cm['within_limit'])}{last_row}"

    RED_FILL   = PatternFill("solid", fgColor="FDECEA")
    GREEN_FILL = PatternFill("solid", fgColor="D4E8C2")
    RED_FONT_D   = Font(name="Arial", color="8B1A1A", bold=True, size=9)
    GREEN_FONT_D = Font(name="Arial", color="00424D", bold=True, size=9)

    for rng in (range_J, range_K):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"{col_J_letter}3>{col_I_letter}3"],
            fill=RED_FILL, font=RED_FONT_D))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f"AND({col_J_letter}3<={col_I_letter}3,{col_J_letter}3>0)"],
            fill=GREEN_FILL, font=GREEN_FONT_D))

    # ── DASHBOARD sheet ───────────────────────────────────────────────────────
    if "DASHBOARD" in wb.sheetnames:
        del wb["DASHBOARD"]
    wd = wb.create_sheet("DASHBOARD", 0)   # first tab

    def dh(hex_): 
        return hex_.lstrip("#")

    BG     = PatternFill("solid", fgColor="00424D")
    HDR    = PatternFill("solid", fgColor="5C9C8E")
    WARN   = PatternFill("solid", fgColor="FDECEA")
    OK     = PatternFill("solid", fgColor="D4E8C2")
    STRIPE = PatternFill("solid", fgColor="F5F5F5")
    PLAIN  = PatternFill("solid", fgColor="FFFFFF")
    bdr    = Border(
        left=Side(style="thin",color="CCCCCC"),
        right=Side(style="thin",color="CCCCCC"),
        top=Side(style="thin",color="CCCCCC"),
        bottom=Side(style="thin",color="CCCCCC"),
    )

    def w(row, col, val, bold=False, size=10, color="000000", bg=None, align="left", wrap=False):
        c = wd.cell(row=row, column=col, value=val)
        c.font      = Font(name="Arial", bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = bdr
        if bg: c.fill = bg
        return c

    # Title
    wd.merge_cells("A1:G1")
    t = wd.cell(row=1, column=1,
                value=f"FORUS TOOLKIT — UPDATE DASHBOARD   (last built {DATE_STAMP}  v{VERSION})")
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="00424D")
    t.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[1].height = 28

    # ── Section 1: OVER LIMIT rows ────────────────────────────────────────────
    wd.merge_cells("A3:G3")
    h1 = wd.cell(row=3, column=1, value="⚠  OVER LIMIT — fix these before distributing")
    h1.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    h1.fill  = PatternFill("solid", fgColor="ED1651")
    h1.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[3].height = 22

    over_hdrs = ["block_id", "section", "block_type", "word_limit", "word_count", "over_by", "content preview"]
    for ci, h in enumerate(over_hdrs, 1):
        c = wd.cell(row=4, column=ci, value=h.upper())
        c.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="C0134A")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
    wd.row_dimensions[4].height = 18

    over_items = [(bid, d) for bid, d in all_rows_by_id.items()
                  if d["word_count"] > d["word_limit"]]
    over_items.sort(key=lambda x: x[1].get("section",""))

    r = 5
    for bid, d in over_items:
        bg = WARN if r % 2 == 0 else PatternFill("solid", fgColor="FFF0EE")
        vals = [bid, d.get("section",""), d.get("block_type",""),
                d["word_limit"], d["word_count"],
                d["word_count"] - d["word_limit"],
                str(d.get("content_text",""))[:60] + "..."]
        for ci, v in enumerate(vals, 1):
            w(r, ci, v, bg=bg)
        wd.row_dimensions[r].height = 18
        r += 1

    if not over_items:
        wd.merge_cells(f"A5:G5")
        c = wd.cell(row=5, column=1, value="✓  No over-limit rows — all content within limits")
        c.font  = Font(name="Arial", bold=True, size=10, color="00424D")
        c.fill  = OK
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
        r = 6

    r += 1  # gap

    # ── Section 2: Review due ─────────────────────────────────────────────────
    wd.merge_cells(f"A{r}:G{r}")
    h2 = wd.cell(row=r, column=1, value="📅  REVIEW DUE — check these are still accurate")
    h2.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    h2.fill  = PatternFill("solid", fgColor="B2C100")
    h2.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[r].height = 22
    r += 1

    rev_hdrs = ["block_id","section","block_type","next_review","update_priority","last_updated","content preview"]
    for ci, h in enumerate(rev_hdrs, 1):
        c = wd.cell(row=r, column=ci, value=h.upper())
        c.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="C9D900")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
    wd.row_dimensions[r].height = 18
    r += 1

    today = datetime.date.today()
    due_items = []
    ws_c = wb["CONTENT"]
    hdrs_c = [c.value for c in ws_c[2]]
    cm_c   = {h: i for i, h in enumerate(hdrs_c) if h}
    for row_data in ws_c.iter_rows(min_row=3, values_only=True):
        bid = row_data[0]
        if not bid: continue
        nr = row_data[cm_c.get("next_review", 14)]
        if nr:
            try:
                nr_date = nr if isinstance(nr, datetime.date) else datetime.date.fromisoformat(str(nr))
                if nr_date <= today:
                    due_items.append({
                        "bid": bid,
                        "section": row_data[cm_c.get("section",2)],
                        "block_type": row_data[cm_c.get("block_type",5)],
                        "next_review": str(nr_date),
                        "update_priority": row_data[cm_c.get("update_priority",11)],
                        "last_updated": str(row_data[cm_c.get("last_updated",12)] or ""),
                        "content": str(row_data[cm_c.get("content_text",7)] or "")[:60],
                    })
            except (ValueError, TypeError):
                pass

    due_items.sort(key=lambda x: x["next_review"])
    for d in due_items:
        bg = PatternFill("solid", fgColor="FFF8DC") if r % 2 == 0 else PLAIN
        vals = [d["bid"], d["section"], d["block_type"], d["next_review"],
                d["update_priority"], d["last_updated"], d["content"]+"..."]
        for ci, v in enumerate(vals, 1):
            w(r, ci, v, bg=bg)
        wd.row_dimensions[r].height = 18
        r += 1

    if not due_items:
        wd.merge_cells(f"A{r}:G{r}")
        c = wd.cell(row=r, column=1, value="✓  No reviews currently due")
        c.font  = Font(name="Arial", bold=True, size=10, color="00424D")
        c.fill  = OK
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
        r += 1

    r += 1  # gap

    # ── Section 3: Content summary ────────────────────────────────────────────
    wd.merge_cells(f"A{r}:G{r}")
    h3 = wd.cell(row=r, column=1, value="📊  CONTENT SUMMARY by section")
    h3.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    h3.fill  = PatternFill("solid", fgColor="00424D")
    h3.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[r].height = 22
    r += 1

    sum_hdrs = ["section","total blocks","steps","tips/warnings","checklists","over limit","notes"]
    for ci, h in enumerate(sum_hdrs, 1):
        c = wd.cell(row=r, column=ci, value=h.upper())
        c.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="58C5C7")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = bdr
    wd.row_dimensions[r].height = 18
    r += 1

    from collections import defaultdict
    by_section = defaultdict(lambda: {"total":0,"steps":0,"callouts":0,"checks":0,"over":0})
    for bid, d in all_rows_by_id.items():
        sec = d.get("section","?")
        by_section[sec]["total"] += 1
        bt = d.get("block_type","")
        if bt == "STEP": by_section[sec]["steps"] += 1
        elif bt in ("TIP","WARNING"): by_section[sec]["callouts"] += 1
        elif bt == "CHECKLIST": by_section[sec]["checks"] += 1
        if d["word_count"] > d["word_limit"]: by_section[sec]["over"] += 1

    for si, (sec, d) in enumerate(sorted(by_section.items())):
        bg = STRIPE if si % 2 == 0 else PLAIN
        over_note = f"⚠ {d['over']} over limit" if d["over"] else "✓ all within limits"
        vals = [sec, d["total"], d["steps"], d["callouts"], d["checks"], d["over"], over_note]
        for ci, v in enumerate(vals, 1):
            c = w(r, ci, v, bg=bg)
            if ci == 7:
                c.font = Font(name="Arial", size=9, bold=True,
                              color="ED1651" if d["over"] else "2D5016")
        wd.row_dimensions[r].height = 18
        r += 1

    # ── Quick reference: word limits ─────────────────────────────────────────
    r += 1
    wd.merge_cells(f"A{r}:G{r}")
    h4 = wd.cell(row=r, column=1,
                 value="📝  WORD LIMITS — default per block type (override in column I of CONTENT sheet)")
    h4.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    h4.fill  = PatternFill("solid", fgColor="4A4A4A")
    h4.alignment = Alignment(horizontal="left", vertical="center")
    wd.row_dimensions[r].height = 22
    r += 1

    limit_pairs = sorted(WORD_LIMITS.items(), key=lambda x: x[1])
    for i in range(0, len(limit_pairs), 4):
        chunk = limit_pairs[i:i+4]
        for ci, (bt, lim) in enumerate(chunk):
            col = ci * 2 + 1
            c1 = wd.cell(row=r, column=col, value=bt)
            c1.font  = Font(name="Arial", bold=True, size=9)
            c1.fill  = PatternFill("solid", fgColor="F0F0F0")
            c1.alignment = Alignment(horizontal="right", vertical="center")
            c1.border = bdr
            c2 = wd.cell(row=r, column=col+1, value=lim)
            c2.font  = Font(name="Arial", size=9, color="1F4E79")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = bdr
        wd.row_dimensions[r].height = 18
        r += 1

    # Column widths for DASHBOARD
    wd.column_dimensions["A"].width = 22
    wd.column_dimensions["B"].width = 28
    wd.column_dimensions["C"].width = 16
    wd.column_dimensions["D"].width = 12
    wd.column_dimensions["E"].width = 12
    wd.column_dimensions["F"].width = 10
    wd.column_dimensions["G"].width = 38

    # ── Hide admin columns in CONTENT sheet ──────────────────────────────────
    # Show only the columns an editor needs; hide tracking/admin columns
    # Visible: A(id) B(part) C(section) E(time_horizon) F(block_type) H(content) I(word_limit) J(word_count) K(within_limit)
    # Hidden:  D(scenario) G(sensitivity) L(update_priority) M(last_updated) N(updated_by) O(next_review) P(language) Q(display_order) R(editor_notes)
    for col_letter in ["D", "G", "L", "M", "N", "O", "P", "Q", "R"]:
        ws.column_dimensions[col_letter].hidden = True

    # Freeze panes so header rows stay visible while scrolling
    ws.freeze_panes = "H3"   # freeze cols A-G and rows 1-2

    # Make content column wider for easier editing
    ws.column_dimensions["H"].width = 70
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10

    wb.save(SPREADSHEET)
    print(f"  ✓ Spreadsheet updated — {len(over_items)} over-limit rows, {len(due_items)} reviews due")


def build_word_count_map(rows):
    """Build {block_id: {word_count, word_limit, section, block_type, content_text}} from all rows."""
    result = {}
    for item in rows:
        bid  = item.get("block_id")
        if not bid: continue
        text = str(item.get("content_text","") or "")
        result[bid] = {
            "word_count":   wc(text),
            "word_limit":   get_limit(item),
            "section":      item.get("section",""),
            "block_type":   item.get("block_type",""),
            "content_text": text,
        }
    return result


def build_pdf(access_level, language="EN"):
    lang  = language.upper() if language else "EN"
    label = "Public" if access_level == 1 else "Network"
    # Use OUT_PUBLIC / OUT_NETWORK exactly as set by the caller.
    # The CLI __main__ block adds language suffixes before calling;
    # the app sets gt.OUT_PUBLIC / gt.OUT_NETWORK to the desired path directly.
    out = OUT_PUBLIC if access_level == 1 else OUT_NETWORK
    print(f"\nBuilding {label} PDF [{lang}] v{VERSION} → {out}")

    rows, mechs = load_data(access_level, language=lang)

    # Full-build req: all parts, all annexes, all regions, all tools included
    full_req = {
        "parts":   {i: True for i in range(1, 8)},
        "annexes": {k: True for k in _ANNEX_CATEGORY},
        "regions": {k: True for k in _REGION_GEO_KEYWORDS},
        "tools":   {"T1": True, "T2": True, "T3": True, "T4": True,
                    "A1": True, "A2": True, "A3": True},
    }

    common_kw = dict(pagesize=A4, leftMargin=ML, rightMargin=MR,
                     topMargin=MT, bottomMargin=MB)

    # ── Pass 1: layout to temp file, collect page→section map ────────────────
    story1, warnings = make_story(rows, mechs, access_level, page_map=None, req=full_req)
    tmp = out.replace(".pdf", "_pass1.pdf")
    doc1 = ToolkitDoc(tmp, access_level=access_level, page_map=None, **common_kw)
    doc1.build(story1)

    # Build page map: page → (part, section) of the FIRST section starting on that page
    page_map = {}
    for (pn, part, section) in doc1._meta_log:
        if pn not in page_map:           # take first (topmost) section per page
            page_map[pn] = (part, section)
    try:
        os.remove(tmp)
    except OSError:
        pass

    # ── Pass 2: build real PDF using page map (ToC gets real page numbers + links) ──
    story2, _ = make_story(rows, mechs, access_level, page_map=page_map, req=full_req)
    # Write to a temp path first so we can append tool pages
    main_tmp = out.replace(".pdf", "_main.pdf")
    doc2 = ToolkitDoc(main_tmp, access_level=access_level, page_map=page_map, **common_kw)
    doc2.build(story2)

    # ── Insert tool pages inline after their section's last page ────────────
    _merge_tools_inline(main_tmp, full_req.get("tools", {}), page_map, out)

    size_kb = os.path.getsize(out) // 1024
    print(f"  ✓ Done — {size_kb}KB — {len(rows)} rows — {out}")
    for w in warnings:
        print(w)



# ── Custom request PDF ────────────────────────────────────────────────────────
# Column positions in REQUEST_LOG (1-indexed, matching spreadsheet order):
_RL_COLS = {
    "req_id":1, "date":2, "name":3, "org":4, "email":5,
    # regions
    "africa":6,"asia":7,"europe":8,"latam":9,"pacific":10,"global":11,
    # parts
    "part1":12,"part2":13,"part3":14,"part4":15,"part5":16,"part6":17,"part7":18,
    # annexes (Part 8 sections)
    "annex_a":19,"annex_b":20,"annex_c":21,
    # admin
    "status":22,"date_sent":23,"by":24,"notes":25,
}

# Maps annex section names to the region column keys that gate their content
_ANNEX_REGIONS = {
    "Annex A: Legal Pro Bono Support":           ["africa","asia","europe","latam","pacific","global"],
    "Annex B: Emergency Grants Mechanisms":       ["africa","asia","europe","latam","pacific","global"],
    "Annex C: Physical & Digital Security Support":["africa","asia","europe","latam","pacific","global"],
}

# Region name → column key mapping (matches SUBSECTION content_text values in spreadsheet)
_REGION_NAME_TO_KEY = {
    "Africa":                    "africa",
    "Asia-Pacific":              "asia",
    "Europe":                    "europe",
    "Latin America & Caribbean": "latam",
    "Pacific":                   "pacific",
    "Global":                    "global",
}

# Annex section name → req_log column key
_ANNEX_TO_KEY = {
    "Annex A: Legal Pro Bono Support":            "annex_a",
    "Annex B: Emergency Grants Mechanisms":        "annex_b",
    "Annex C: Physical & Digital Security Support":"annex_c",
}

# Annex section name → ANNEXES sheet category value
_ANNEX_CATEGORY = {
    "Annex A: Legal Pro Bono Support":             "legal",
    "Annex B: Emergency Grants Mechanisms":        "emergency-funding",
    "Annex C: Physical & Digital Security Support":"digital-security",
}

# Region key → keywords to look for in geographic_coverage field
_REGION_GEO_KEYWORDS = {
    "africa":  ["Africa", "Sub-Saharan"],
    "asia":    ["Asia"],
    "europe":  ["Europe"],
    "latam":   ["Americas", "Latin America", "Caribbean"],
    "pacific": ["Pacific"],
    "global":  ["Global"],
}
_GEO_ALL_REGIONS = ["all region", "All region", "all regions", "All regions"]


def read_request(req_id):
    """Read a single request row from REQUEST_LOG. Returns dict or None."""
    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    if "REQUEST_LOG" not in wb.sheetnames:
        print(f"ERROR: REQUEST_LOG sheet not found in {SPREADSHEET}")
        return None
    ws = wb["REQUEST_LOG"]
    # Headers are on row 4; data starts row 5
    col = _RL_COLS
    for row in ws.iter_rows(min_row=5, values_only=True):
        if str(row[col["req_id"]-1] or "").strip() == req_id:
            def yn(key):
                v = str(row[col[key]-1] or "").strip().upper()
                return v == "Y"
            return {
                "req_id":  req_id,
                "name":    str(row[col["name"]-1] or ""),
                "org":     str(row[col["org"]-1] or ""),
                "email":   str(row[col["email"]-1] or ""),
                # regions
                "regions": {k: yn(k) for k in
                    ["africa","asia","europe","latam","pacific","global"]},
                # parts 1-7
                "parts":   {i+1: yn(f"part{i+1}") for i in range(7)},
                # annexes
                "annexes": {
                    "Annex A: Legal Pro Bono Support":            yn("annex_a"),
                    "Annex B: Emergency Grants Mechanisms":        yn("annex_b"),
                    "Annex C: Physical & Digital Security Support":yn("annex_c"),
                },
            }
    print(f"ERROR: Request {req_id!r} not found in REQUEST_LOG")
    return None


def filter_rows_for_request(rows, req):
    """Filter content rows to only those selected in the request."""
    selected_parts = {p for p, v in req["parts"].items() if v}
    selected_annex_secs = {sec for sec, v in req["annexes"].items() if v}
    selected_regions = {k for k, v in req["regions"].items() if v}

    out = []
    for item in rows:
        part    = int(item.get("part",1) or 1)
        section = str(item.get("section",""))
        btype   = str(item.get("block_type","")).strip()

        # Part 1–7: include only if that part is ticked
        if part <= 7:
            if part not in selected_parts:
                continue
            out.append(item)
            continue

        # Part 8 annexes: include section only if annex is ticked
        if part == 8:
            # Which annex does this row belong to?
            if section not in selected_annex_secs:
                continue
            # Within included annexes, filter by region
            # SUBSECTION and COUNTRY-ENTRY rows are region-specific
            if btype in ("SUBSECTION", "COUNTRY-ENTRY"):
                text = str(item.get("content_text",""))
                if "|" in text:
                    country_or_region = text.split("|",1)[0].strip()
                else:
                    country_or_region = text.strip()
                # Match the region name against selected regions.
                # Only include rows whose region tag is in the selected set.
                # "Global" is treated as a region like any other — ticking it
                # shows globally-applicable entries, not all regional entries.
                region_key = _REGION_NAME_TO_KEY.get(country_or_region)
                if region_key and region_key not in selected_regions:
                    continue
            out.append(item)

    return out


def build_request_pdf(req_id, access_level=1):
    """Generate a single customised PDF for a specific request row."""
    req = read_request(req_id)
    if not req:
        return

    label  = "Public" if access_level == 1 else "Network"
    suffix = f"_{req_id}"
    out    = OUT_PUBLIC.replace(".pdf", f"{suffix}.pdf") if access_level == 1              else OUT_NETWORK.replace(".pdf", f"{suffix}.pdf")

    print(f"\nBuilding custom PDF for {req_id} ({req['name']}, {req['org']})")
    selected_parts = [p for p,v in req["parts"].items() if v]
    selected_annex = [s.split(":")[0] for s,v in req["annexes"].items() if v]
    print(f"  Parts: {selected_parts}  |  Annexes: {selected_annex}")
    print(f"  Regions: {[k for k,v in req['regions'].items() if v]}")
    print(f"  Output: {out}")

    all_rows, mechs = load_data(access_level)
    rows = filter_rows_for_request(all_rows, req)

    common_kw = dict(pagesize=A4, leftMargin=ML, rightMargin=MR,
                     topMargin=MT, bottomMargin=MB)

    # Pass 1
    story1, warnings = make_story(rows, mechs, access_level, page_map=None, req=req)
    tmp = out.replace(".pdf","_pass1.pdf")
    doc1 = ToolkitDoc(tmp, access_level=access_level, page_map=None, **common_kw)
    doc1.build(story1)
    page_map = {}
    for (pn, part, section) in doc1._meta_log:
        if pn not in page_map:
            page_map[pn] = (part, section)
    try: os.remove(tmp)
    except OSError: pass

    # Pass 2
    story2, _ = make_story(rows, mechs, access_level, page_map=page_map, req=req)
    doc2 = ToolkitDoc(out, access_level=access_level, page_map=page_map, **common_kw)
    doc2.build(story2)

    size_kb = os.path.getsize(out) // 1024
    print(f"  ✓ Done — {size_kb}KB — {len(rows)} rows — {out}")
    for w in warnings:
        print(w)

    # Mark as generated in REQUEST_LOG
    _mark_request_sent(req_id)


def build_pdf_from_request_dict(req, access_level=1, out_path=None, language="EN"):
    """Generate a customised PDF directly from a request dict (no REQUEST_LOG lookup).

    req must contain:
        name, org, email (str)
        parts   : {1: bool, 2: bool, … 7: bool}
        regions : {africa: bool, asia: bool, europe: bool, latam: bool, pacific: bool, global: bool}
        annexes : {"Annex A: Legal Pro Bono Support": bool, …}
    Returns True on success, False on failure.
    """
    if not req:
        return False

    lang  = language.upper() if language else "EN"
    label = "Public" if access_level == 1 else "Network"
    if out_path is None:
        suffix = f"_{req.get('org','custom').replace(' ','_')}"
        out_path = (OUT_PUBLIC.replace(".pdf", f"{suffix}.pdf") if access_level == 1
                    else OUT_NETWORK.replace(".pdf", f"{suffix}.pdf"))

    print(f"\nBuilding custom PDF for {req.get('name','')}, {req.get('org','')} [{lang}]")

    all_rows, mechs = load_data(access_level, language=lang)
    rows = filter_rows_for_request(all_rows, req)

    common_kw = dict(pagesize=A4, leftMargin=ML, rightMargin=MR,
                     topMargin=MT, bottomMargin=MB)

    # Pass 1
    story1, warnings = make_story(rows, mechs, access_level, page_map=None, req=req)
    tmp = out_path.replace(".pdf", "_pass1.pdf")
    doc1 = ToolkitDoc(tmp, access_level=access_level, page_map=None, **common_kw)
    doc1.build(story1)
    page_map = {}
    for (pn, part, section) in doc1._meta_log:
        if pn not in page_map:
            page_map[pn] = (part, section)
    try:
        os.remove(tmp)
    except OSError:
        pass

    # Pass 2
    story2, _ = make_story(rows, mechs, access_level, page_map=page_map, req=req)
    main_tmp = out_path.replace(".pdf", "_main.pdf")
    doc2 = ToolkitDoc(main_tmp, access_level=access_level, page_map=page_map, **common_kw)
    doc2.build(story2)

    # ── Insert tool pages inline after their section's last page ────────────
    _merge_tools_inline(main_tmp, req.get("tools", {}), page_map, out_path)

    if os.path.exists(out_path):
        size_kb = os.path.getsize(out_path) // 1024
        print(f"  ✓ Done — {size_kb}KB — {len(rows)} rows — {out_path}")
        for w in warnings:
            print(w)
        return True

    return False


def _mark_request_sent(req_id):
    """Update STATUS to IN PROGRESS in REQUEST_LOG after PDF is generated."""
    wb = openpyxl.load_workbook(SPREADSHEET)
    ws = wb["REQUEST_LOG"]
    status_col = _RL_COLS["status"]
    for row in ws.iter_rows(min_row=5):
        if str(row[0].value or "").strip() == req_id:
            status_cell = row[status_col - 1]
            if str(status_cell.value or "").strip() == "PENDING":
                status_cell.value = "IN PROGRESS"
            break
    wb.save(SPREADSHEET)




# ── Auto-update: constants ───────────────────────────────────────────────────

# Column order of REVIEW_QUEUE sheet (matches spreadsheet header row)
_RQ_HDRS = [
    "approve", "review_id", "date_flagged", "mech_id", "mechanism_name",
    "category", "change_type", "field", "current_value", "proposed_value",
    "reason", "source_url", "confidence", "status",
    "reviewer_notes", "reviewed_by", "reviewed_date",
]

# Default verification interval in months by category
_VERIFY_MONTHS = {
    "legal":             1,
    "emergency-funding": 1,
    "digital-security":  1,
}

# ── Category-specific research prompts ───────────────────────────────────────
_RESEARCH_PROMPTS = {

"legal": """You are verifying a legal pro bono support entry in the Forus Resilience & Support Toolkit.

The readers of this entry are national civil society platforms and coalitions (Forus network members) who need to quickly decide whether this mechanism is accessible and relevant to them. They do not need organisational background, history, network size statistics, or any detail that does not directly help them decide whether to apply and how to do so.

Check specifically:
- Is the organisation still operating and providing pro bono legal support to civil society?
- Has the referral or intake process changed (e.g. new online form, different contact route)?
- Have thematic focus areas expanded or narrowed?
- Has geographic coverage changed?
- Is the contact information (email, URL) still valid and resolving?
- Has platform eligibility changed (do they now serve or no longer serve NGO networks/platforms)?

Typical sources: organisation's own website, annual reports, ICNL civic freedom resources, PILnet directory.

IMPORTANT — proposed_value writing rules:
1. Length: proposed_value must be no longer than the current_value it replaces. Revise in place — update the specific facts that have changed. Do not add new sentences or detail not already in the current value.
2. Relevance: include only information a national platform needs to decide whether to apply and how. Omit organisational history, network size figures, internal statistics, and any background that does not affect access or eligibility.

If the entry is fully accurate, return status NO_CHANGE.
If you find changes, list each changed field with proposed new text and the source URL.
If you cannot verify, return status UNABLE_TO_VERIFY with a note explaining what was searched.""",

"emergency-funding": """You are verifying an emergency grant mechanism entry in the Forus Resilience & Support Toolkit.

The readers of this entry are national civil society platforms and coalitions (Forus network members) who need to quickly decide whether this fund is open and relevant to them. They do not need fund history, donor relationships, total disbursement figures, or any detail that does not directly help them decide whether to apply and how.

Check specifically:
- Is the fund still accepting applications?
- Has the grant size range changed?
- Have eligibility criteria changed (geographic, thematic, organisational type)?
- Has the application process or timeline changed?
- Has the typical response/disbursement timeframe changed?
- Has the fund been closed, paused, or merged with another mechanism?
- Are there new emergency funding mechanisms that should be added? (include in new_mechanisms_found)

Typical sources: fund's own website, Fundsforngos.org, CIVICUS Monitor, civic space news sources.

IMPORTANT — proposed_value writing rules:
1. Length: proposed_value must be no longer than the current_value it replaces. Revise in place — update the specific facts that have changed. Do not add new sentences or detail not already in the current value.
2. Relevance: include only information a national platform needs to decide whether to apply and how. Omit fund history, donor lists, total disbursement figures, and any background that does not affect access or eligibility.

If the entry is fully accurate, return status NO_CHANGE.
If you find changes, list each changed field with proposed new text and the source URL.
If you cannot verify, return status UNABLE_TO_VERIFY with a note.""",

"digital-security": """You are verifying a digital and physical security support entry in the Forus Resilience & Support Toolkit.

The readers of this entry are national civil society platforms and coalitions (Forus network members) who need to quickly decide whether this support service is accessible and relevant to them. They do not need provider history, case statistics, or any detail that does not directly help them decide whether to reach out and how.

Check specifically:
- Is the helpline or service still operational?
- Has language coverage changed?
- Have response timeframes changed?
- Has the secure contact method changed (e.g. new PGP key, new Signal number, new encrypted form)?
- Has platform/network eligibility changed (as distinct from individual eligibility)?
- Are there new security support providers that should be added? (include in new_mechanisms_found)

Typical sources: provider's own website, Access Now reports, Digital Defenders Partnership, CiviCERT.

IMPORTANT — proposed_value writing rules:
1. Length: proposed_value must be no longer than the current_value it replaces. Revise in place — update the specific facts that have changed. Do not add new sentences or detail not already in the current value.
2. Relevance: include only information a national platform needs to decide whether to reach out and how. Omit provider history, case volume statistics, and any background that does not affect access or eligibility.

If the entry is fully accurate, return status NO_CHANGE.
If you find changes, list each changed field with proposed new text and the source URL.
If you cannot verify, return status UNABLE_TO_VERIFY with a note.""",
}


# ── Auto-update: helper functions ────────────────────────────────────────────

def _rq_append(ws_rq, item_dict, next_id_holder):
    """Append one row to the REVIEW_QUEUE sheet and return its new review_id."""
    next_id_holder[0] += 1
    rid = f"RQ-{next_id_holder[0]:04d}"
    ws_rq.append([
        None,                                       # approve (checkbox — left blank)
        rid,                                        # review_id
        item_dict.get("date_flagged", ""),
        item_dict.get("mech_id", ""),
        item_dict.get("mechanism_name", ""),
        item_dict.get("category", ""),
        item_dict.get("change_type", "UPDATED_INFO"),
        item_dict.get("field", ""),
        item_dict.get("current_value", ""),
        item_dict.get("proposed_value", ""),
        item_dict.get("reason", ""),
        item_dict.get("source_url", ""),
        item_dict.get("confidence", "MEDIUM"),
        "PENDING",
        None,   # reviewer_notes
        None,   # reviewed_by
        None,   # reviewed_date
    ])
    return rid


def _agent_error(mech_id, msg):
    return {
        "mech_id":              mech_id,
        "status":               "UNABLE_TO_VERIFY",
        "confidence":           "LOW",
        "changes":              [],
        "new_mechanisms_found": [],
        "notes":                msg,
    }


def call_ai_agent(mech_dict, api_key):
    """Call the Anthropic API to research one mechanism row.
    Returns a parsed dict matching the PRD JSON output schema."""
    import json, urllib.request, urllib.error

    mech_id  = str(mech_dict.get("mech_id", ""))
    category = str(mech_dict.get("category", "legal")).strip()
    prompt   = _RESEARCH_PROMPTS.get(category, _RESEARCH_PROMPTS["legal"])

    # Serialise the row for the prompt, skipping empty values
    row_lines = [f"  {k}: {v}" for k, v in mech_dict.items()
                 if v is not None and str(v).strip()]
    mech_context = "\n".join(row_lines)

    user_msg = (
        f"Current entry data:\n{mech_context}\n\n"
        f"{prompt.strip()}\n\n"
        "Return ONLY a valid JSON object — no prose before or after it.\n"
        "Reminder: every proposed_value must be no longer than the current_value it replaces, and must contain "
        "only information a national civil society platform needs to decide whether to apply and how. "
        "Omit organisational history, statistics, and any background that does not affect access or eligibility. "
        "Update facts in place; do not expand or enrich the text.\n"
        "{\n"
        f"  \"mech_id\": \"{mech_id}\",\n"
        "  \"status\": \"CHANGE_DETECTED\" | \"NO_CHANGE\" | \"UNABLE_TO_VERIFY\",\n"
        "  \"confidence\": \"HIGH\" | \"MEDIUM\" | \"LOW\",\n"
        "  \"changes\": [\n"
        "    {\n"
        "      \"field\": \"field_name\",\n"
        "      \"current_value\": \"...\",\n"
        "      \"proposed_value\": \"...\",\n"
        "      \"reason\": \"...\",\n"
        "      \"source_url\": \"...\"\n"
        "    }\n"
        "  ],\n"
        "  \"new_mechanisms_found\": [\n"
        "    { \"name\": \"...\", \"organisation\": \"...\", \"url\": \"...\", \"reason\": \"...\" }\n"
        "  ],\n"
        "  \"notes\": \"...\"\n"
        "}"
    )

    payload = json.dumps({
        "model":      "claude-sonnet-4-6",
        "max_tokens": 2048,
        "tools":      [{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
        "messages":   [{"role": "user", "content": user_msg}],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type":       "application/json",
            "x-api-key":          api_key,
            "anthropic-version":  "2023-06-01",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")[:300]
        return _agent_error(mech_id, f"HTTP {e.code}: {body}")
    except Exception as e:
        return _agent_error(mech_id, str(e))

    # Extract text blocks from the response
    text = "".join(
        block.get("text", "") for block in raw.get("content", [])
        if block.get("type") == "text"
    )

    # Locate the outermost JSON object
    start = text.find("{")
    end   = text.rfind("}") + 1
    if start < 0 or end <= start:
        return _agent_error(mech_id, f"No JSON in response: {text[:200]}")

    json_str = text[start:end]
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        pass

    # Fallback: strip control characters and retry
    import re as _re
    cleaned = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', json_str)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError as e2:
        # Last resort: try to extract just the structural fields we need
        # by asking the API again with a stricter instruction
        try:
            status_m  = _re.search(r'"status"\s*:\s*"([^"]+)"', json_str)
            conf_m    = _re.search(r'"confidence"\s*:\s*"([^"]+)"', json_str)
            # If we can at least get status, return a minimal valid result
            if status_m:
                return {
                    "mech_id":    mech_id,
                    "status":     status_m.group(1),
                    "confidence": conf_m.group(1) if conf_m else "LOW",
                    "changes":    [],
                    "new_mechanisms_found": [],
                    "notes":      "JSON partially recovered — changes list could not be parsed",
                }
        except Exception:
            pass
        return _agent_error(mech_id, f"JSON parse error: {e2}")


def _translate_field(en_value, field, lang, api_key):
    """Translate a single ANNEX field value from English to FR or ES.

    Returns the translated string, or None on failure.
    Preserves proper nouns, URLs, email addresses, and mechanism IDs.
    """
    import json, urllib.request, urllib.error

    lang_name = {"FR": "French", "ES": "Spanish"}.get(lang.upper())
    if not lang_name or not en_value or not api_key:
        return None

    user_msg = (
        f"Translate the following text from English into {lang_name}.\n\n"
        "Rules:\n"
        "1. Preserve all proper nouns (organisation names, fund names, programme names).\n"
        "2. Preserve all URLs, email addresses, and mechanism IDs exactly as written.\n"
        "3. Preserve currency codes, amounts, and technical terms.\n"
        "4. Keep the same sentence length and register as the original.\n"
        "5. Return ONLY the translated text — no explanation, no quotes around it.\n\n"
        f"Field: {field}\n"
        f"Text to translate:\n{en_value}"
    )

    payload = json.dumps({
        "model":      "claude-haiku-4-5-20251001",
        "max_tokens": 1024,
        "messages":   [{"role": "user", "content": user_msg}],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type":      "application/json",
            "x-api-key":         api_key,
            "anthropic-version": "2023-06-01",
        },
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
        text_blocks = [b.get("text", "") for b in raw.get("content", []) if b.get("type") == "text"]
        return "".join(text_blocks).strip() or None
    except Exception:
        return None


def _update_mech_verified(ws_m, row_idx, mcm, today, category):
    """Update last_verified, verified_by, next_verify_due for a NO_CHANGE row."""
    import datetime
    months = _VERIFY_MONTHS.get(str(category).strip(), 12)
    nxt = today.replace(
        month=((today.month - 1 + months) % 12) + 1,
        year=today.year + ((today.month - 1 + months) // 12),
    )
    ws_m.cell(row=row_idx, column=mcm.get("last_verified",  14) + 1).value = str(today)
    ws_m.cell(row=row_idx, column=mcm.get("verified_by",    15) + 1).value = "AI Agent"
    ws_m.cell(row=row_idx, column=mcm.get("next_verify_due",16) + 1).value = str(nxt)


# ── Auto-update: main entry points ───────────────────────────────────────────

def check_mechanisms(api_key):
    """--check-mechanisms mode.
    Reads ANNEXES rows due for verification, calls the AI agent for each,
    and writes proposals to the REVIEW_QUEUE sheet.
    NO_CHANGE results update last_verified/next_verify_due directly."""
    import datetime

    if not api_key:
        print("ERROR: Anthropic API key required.")
        print("  Pass --api-key KEY  or  set env var ANTHROPIC_API_KEY.")
        return

    wb   = openpyxl.load_workbook(SPREADSHEET)
    ws_m = wb["ANNEXES"]

    mhdrs = [c.value for c in ws_m[2]]
    mcm   = {h: i for i, h in enumerate(mhdrs) if h}   # name → 0-indexed

    today  = datetime.date.today()
    window = today + datetime.timedelta(days=14)

    # Collect due rows
    due_rows = []
    for row_idx, row in enumerate(ws_m.iter_rows(min_row=3, values_only=False), start=3):
        mech_id = row[0].value
        if not mech_id:
            continue
        status_val = str(row[mcm.get("status", 4)].value or "").strip().upper()
        nvd = row[mcm.get("next_verify_due", 16)].value
        if nvd:
            if isinstance(nvd, datetime.datetime):
                nvd = nvd.date()
            elif not isinstance(nvd, datetime.date):
                try:
                    nvd = datetime.date.fromisoformat(str(nvd))
                except (ValueError, TypeError):
                    nvd = None
        is_due = (nvd and nvd <= window) or (status_val == "VERIFY")
        if is_due:
            mech_dict = {mhdrs[i]: row[i].value for i in range(len(mhdrs)) if i < len(row)}
            due_rows.append((row_idx, mech_dict))

    if not due_rows:
        print("  ✓ No mechanisms are currently due for verification.")
        return

    print(f"  {len(due_rows)} mechanism(s) due for verification:")
    for _, m in due_rows:
        print(f"    {m['mech_id']}  {m.get('mechanism_name','')}  [{m.get('category','')}]")
    print()

    # Find current max RQ ID so we can generate sequential IDs
    ws_rq = wb["REVIEW_QUEUE"]
    max_n = 0
    for row in ws_rq.iter_rows(min_row=3, values_only=True):
        rid = str(row[1] or "")
        if rid.startswith("RQ-"):
            try:
                max_n = max(max_n, int(rid[3:]))
            except ValueError:
                pass
    next_id = [max_n]   # mutable holder incremented by _rq_append

    # Check each due mechanism
    checked = no_change = proposed = unable = 0

    for row_idx, mech_dict in due_rows:
        mech_id = mech_dict["mech_id"]
        name    = mech_dict.get("mechanism_name", "")
        cat     = mech_dict.get("category", "")
        print(f"  Checking {mech_id} ({name})…", end="", flush=True)

        # Strip translated columns before sending to AI — it should only review EN fields
        en_only_dict = {k: v for k, v in mech_dict.items()
                        if not (k.endswith("_fr") or k.endswith("_es"))}
        result  = call_ai_agent(en_only_dict, api_key)
        status  = result.get("status", "UNABLE_TO_VERIFY")
        conf    = result.get("confidence", "LOW")
        checked += 1

        if status == "NO_CHANGE":
            no_change += 1
            print(f" ✓ No change  [{conf}]")
            _update_mech_verified(ws_m, row_idx, mcm, today, cat)

        elif status == "CHANGE_DETECTED":
            # Filter out any _fr / _es fields the AI may have proposed anyway
            changes = [ch for ch in result.get("changes", [])
                       if not (str(ch.get("field","")).endswith("_fr")
                               or str(ch.get("field","")).endswith("_es"))]
            proposed += len(changes)
            print(f" ⚠  {len(changes)} change(s) detected  [{conf}]")
            for ch in changes:
                _rq_append(ws_rq, {
                    "date_flagged":   str(today),
                    "mech_id":        mech_id,
                    "mechanism_name": name,
                    "category":       cat,
                    "change_type":    "UPDATED_INFO",
                    "field":          ch.get("field", ""),
                    "current_value":  ch.get("current_value", ""),
                    "proposed_value": ch.get("proposed_value", ""),
                    "reason":         ch.get("reason", ""),
                    "source_url":     ch.get("source_url", ""),
                    "confidence":     conf,
                }, next_id)
            # Any newly discovered mechanisms go in as NEW_ENTRY items
            for nm in result.get("new_mechanisms_found", []):
                proposed += 1
                _rq_append(ws_rq, {
                    "date_flagged":   str(today),
                    "mech_id":        mech_id,
                    "mechanism_name": nm.get("name", ""),
                    "category":       cat,
                    "change_type":    "NEW_ENTRY",
                    "field":          "new_mechanism",
                    "current_value":  "",
                    "proposed_value": (f"{nm.get('name','')}  |  "
                                       f"{nm.get('organisation','')}  |  "
                                       f"{nm.get('url','')}"),
                    "reason":         nm.get("reason", ""),
                    "source_url":     nm.get("url", ""),
                    "confidence":     conf,
                }, next_id)

        else:   # UNABLE_TO_VERIFY
            unable += 1
            notes = result.get("notes", "")
            print(f" ? Unable to verify  [{conf}]  {notes[:60]}")
            _rq_append(ws_rq, {
                "date_flagged":   str(today),
                "mech_id":        mech_id,
                "mechanism_name": name,
                "category":       cat,
                "change_type":    "UNABLE_TO_VERIFY",
                "field":          "all",
                "current_value":  "",
                "proposed_value": "",
                "reason":         notes,
                "source_url":     "",
                "confidence":     conf,
            }, next_id)

    wb.save(SPREADSHEET)
    print()
    print(f"  ✓ {checked} checked — {no_change} unchanged, "
          f"{proposed} proposal(s) added to REVIEW_QUEUE, {unable} unable to verify")
    print("  → Open REVIEW_QUEUE tab in the spreadsheet to review and approve.")


def show_review_queue():
    """--review mode. Print a readable terminal summary of all PENDING items."""
    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    if "REVIEW_QUEUE" not in wb.sheetnames:
        print("  No REVIEW_QUEUE sheet found.")
        return
    ws_rq = wb["REVIEW_QUEUE"]
    hdrs  = [c.value for c in ws_rq[2]]
    cm    = {h: i for i, h in enumerate(hdrs) if h}

    pending = []
    for row in ws_rq.iter_rows(min_row=3, values_only=True):
        if not row[cm.get("review_id", 1)]:
            continue
        if str(row[cm.get("status", 13)] or "").strip().upper() == "PENDING":
            pending.append(row)

    if not pending:
        print("  ✓ No pending items in the REVIEW_QUEUE.")
        return

    print(f"  {len(pending)} item(s) pending review:\n")
    rule = "  " + "─" * 58
    for row in pending:
        rid    = row[cm.get("review_id",       1)]
        mech   = row[cm.get("mechanism_name",  4)]
        cat    = row[cm.get("category",        5)]
        ctyp   = row[cm.get("change_type",     6)]
        fld    = row[cm.get("field",           7)]
        conf   = row[cm.get("confidence",     12)]
        cur    = str(row[cm.get("current_value",   8)] or "")
        prop   = str(row[cm.get("proposed_value",  9)] or "")
        reason = str(row[cm.get("reason",         10)] or "")
        src    = row[cm.get("source_url",         11)]
        print(rule)
        print(f"  {rid}  [{conf}]  {ctyp}")
        print(f"  Mechanism : {mech} ({cat})")
        print(f"  Field     : {fld}")
        print(f"  Current   : {cur[:100]}{'…' if len(cur) > 100 else ''}")
        print(f"  Proposed  : {prop[:100]}{'…' if len(prop) > 100 else ''}")
        print(f"  Reason    : {reason[:120]}")
        if src:
            print(f"  Source    : {src}")
    print(rule)
    print("\n  To approve: set Status → APPROVED in the REVIEW_QUEUE tab,")
    print("  then run:   python generate_toolkit.py --apply-approved")


def apply_approved(reviewer_name=None, api_key=None):
    """--apply-approved mode.
    Reads APPROVED items from REVIEW_QUEUE, writes new values back to
    ANNEXES, updates verification dates, and marks items COMPLETED.
    If api_key is provided, auto-translates changed EN fields into FR and ES."""
    import datetime

    wb    = openpyxl.load_workbook(SPREADSHEET)
    ws_rq = wb["REVIEW_QUEUE"]
    ws_m  = wb["ANNEXES"]

    mhdrs = [c.value for c in ws_m[2]]
    mcm   = {h: i for i, h in enumerate(mhdrs) if h}   # field name → 0-indexed col

    # Build mech_id → spreadsheet row number index
    mech_row_map = {}
    for r_idx, row in enumerate(ws_m.iter_rows(min_row=3, values_only=False), start=3):
        mid = row[0].value
        if mid:
            mech_row_map[str(mid).strip()] = r_idx

    rq_hdrs = [c.value for c in ws_rq[2]]
    rqcm    = {h: i for i, h in enumerate(rq_hdrs) if h}   # 0-indexed

    today    = datetime.date.today()
    reviewer = reviewer_name or "AI Agent (approved)"
    applied  = 0
    skipped  = 0

    for r_idx, row in enumerate(ws_rq.iter_rows(min_row=3, values_only=False), start=3):
        rid_cell    = row[rqcm.get("review_id",   1)]
        status_cell = row[rqcm.get("status",     13)]
        if not rid_cell.value:
            continue
        if str(status_cell.value or "").strip().upper() != "APPROVED":
            continue

        mech_id      = str(row[rqcm.get("mech_id",       3)].value or "").strip()
        field        = str(row[rqcm.get("field",          7)].value or "").strip()
        proposed_val = row[rqcm.get("proposed_value",     9)].value
        cat          = str(row[rqcm.get("category",       5)].value or "").strip()

        if field == "new_mechanism":
            # New entries need manual addition; flag and move on
            status_cell.value = "COMPLETED_MANUAL"
            print(f"  {rid_cell.value}: New entry — add manually to ANNEXES sheet.")
            applied += 1
            continue

        mech_row_idx = mech_row_map.get(mech_id)
        if not mech_row_idx:
            print(f"  {rid_cell.value}: WARNING — mech_id {mech_id!r} not found. Skipping.")
            skipped += 1
            continue

        field_col = mcm.get(field)
        if field_col is None:
            print(f"  {rid_cell.value}: WARNING — field {field!r} not in schema. Skipping.")
            skipped += 1
            continue

        # Write proposed value to ANNEXES (EN)
        ws_m.cell(row=mech_row_idx, column=field_col + 1).value = proposed_val

        # Auto-translate changed field into FR and ES if api_key available
        _translatable = ("eligibility_note", "how_to_access", "timeframe", "constraints", "notes")
        if api_key and field in _translatable and proposed_val:
            for lang in ("FR", "ES"):
                translated = _translate_field(str(proposed_val), field, lang, api_key)
                if translated:
                    tr_col = mcm.get(f"{field}_{lang.lower()}")
                    if tr_col is not None:
                        ws_m.cell(row=mech_row_idx, column=tr_col + 1).value = translated
                        print(f"    ↳ Translated {field} → {lang}")
                    else:
                        print(f"    ↳ WARNING: no column {field}_{lang.lower()} in schema")
                else:
                    print(f"    ↳ Translation to {lang} failed — EN value kept")

        # Update verification metadata
        months = _VERIFY_MONTHS.get(cat, 12)
        nxt = today.replace(
            month=((today.month - 1 + months) % 12) + 1,
            year=today.year + ((today.month - 1 + months) // 12),
        )
        ws_m.cell(row=mech_row_idx, column=mcm.get("last_verified",  14) + 1).value = str(today)
        ws_m.cell(row=mech_row_idx, column=mcm.get("verified_by",    15) + 1).value = reviewer
        ws_m.cell(row=mech_row_idx, column=mcm.get("next_verify_due",16) + 1).value = str(nxt)

        # Mark REVIEW_QUEUE item completed
        status_cell.value                          = "COMPLETED"
        row[rqcm.get("reviewed_by",   15)].value  = reviewer
        row[rqcm.get("reviewed_date", 16)].value  = str(today)

        print(f"  {rid_cell.value}: ✓ Applied  {mech_id}.{field}")
        applied += 1

    wb.save(SPREADSHEET)
    print(f"\n  ✓ {applied} item(s) applied, {skipped} skipped.")
    if applied:
        print("  → Regenerate the PDF:  python generate_toolkit.py")


if __name__ == "__main__":
    print("=" * 60)
    print(f"Forus Toolkit PDF Generator  —  v{VERSION}")
    print(f"Date: {DATE_STAMP}  |  Source: {SPREADSHEET}")
    print("=" * 60)
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: {SPREADSHEET} not found.")
        sys.exit(1)

    # ── Auto-update: check mechanisms ────────────────────────────────────────
    if "--check-mechanisms" in sys.argv:
        import os as _os
        api_key = None
        if "--api-key" in sys.argv:
            idx = sys.argv.index("--api-key")
            if idx + 1 < len(sys.argv):
                api_key = sys.argv[idx + 1]
        if not api_key:
            api_key = _os.environ.get("ANTHROPIC_API_KEY")
        print("\nChecking mechanisms for updates…")
        check_mechanisms(api_key)
        print("\n" + "=" * 60)
        sys.exit(0)

    # ── Auto-update: show review queue ───────────────────────────────────────
    if "--review" in sys.argv:
        print("\nReview Queue — Pending Items")
        print("=" * 60)
        show_review_queue()
        print("\n" + "=" * 60)
        sys.exit(0)

    # ── Auto-update: apply approved changes ──────────────────────────────────
    if "--apply-approved" in sys.argv:
        reviewer = None
        if "--reviewer" in sys.argv:
            idx = sys.argv.index("--reviewer")
            if idx + 1 < len(sys.argv):
                reviewer = sys.argv[idx + 1]
        print("\nApplying approved changes…")
        apply_approved(reviewer_name=reviewer)
        print("\n" + "=" * 60)
        sys.exit(0)

    # ── Custom request mode: python generate_toolkit.py --request REQ-001 ──
    if "--request" in sys.argv:
        idx = sys.argv.index("--request")
        if idx + 1 >= len(sys.argv):
            print("ERROR: --request requires a REQ_ID argument, e.g. --request REQ-001")
            sys.exit(1)
        req_id = sys.argv[idx + 1]
        access = 1  # default public; add --network flag to get sensitivity-2 rows
        if "--network" in sys.argv:
            access = 2
        build_request_pdf(req_id, access_level=access)
        print("\n" + "=" * 60)
        print(f"Custom PDF generated for {req_id}.")
        print("  → Mark as SENT in REQUEST_LOG once delivered to the member.")
        print("=" * 60)
        sys.exit(0)

    # ── Language selection ────────────────────────────────────────────────────
    language = "EN"
    if "--language" in sys.argv:
        idx = sys.argv.index("--language")
        if idx + 1 < len(sys.argv):
            language = sys.argv[idx + 1].upper()
    if language not in ("EN", "FR", "ES"):
        print(f"ERROR: --language must be EN, FR, or ES (got '{language}')")
        sys.exit(1)

    # ── Standard full build ───────────────────────────────────────────────────
    # For non-EN CLI builds, suffix the output filenames before calling build_pdf
    global OUT_PUBLIC, OUT_NETWORK
    if language != "EN":
        OUT_PUBLIC  = OUT_PUBLIC.replace(".pdf",  f"_{language}.pdf")
        OUT_NETWORK = OUT_NETWORK.replace(".pdf", f"_{language}.pdf")

    build_pdf(1, language=language)
    build_pdf(2, language=language)

    # Write actual word counts back to spreadsheet and rebuild DASHBOARD tab
    # (word counts are always computed from EN rows)
    all_rows, _ = load_data(2, language="EN")
    count_map = build_word_count_map(all_rows)
    print("\nUpdating spreadsheet (word counts + dashboard)...")
    update_spreadsheet(count_map)

    print("\n" + "=" * 60)
    print("Done.")
    print("  → Standard build:       python generate_toolkit.py")
    print("  → French build:         python generate_toolkit.py --language FR")
    print("  → Spanish build:        python generate_toolkit.py --language ES")
    print("  → Custom member PDF:    python generate_toolkit.py --request REQ-001")
    print("  → Check mechanisms:     python generate_toolkit.py --check-mechanisms [--api-key KEY]")
    print("  → Review queue:         python generate_toolkit.py --review")
    print("  → Apply approved:       python generate_toolkit.py --apply-approved [--reviewer NAME]")
    print("  → To set a custom word limit for any row: type a number in column I.")
    print("=" * 60)
