"""
app.py  —  Forus Toolkit Manager
Streamlit web interface for the Forus Resilience & Support Toolkit.

Run locally:   streamlit run app.py
Deploy to:     Streamlit Community Cloud (connect GitHub repo, set secrets)

Secrets required:
  APP_PASSWORD      — Secretariat (admin) login password
  MEMBER_PASSWORD   — Forus member login password
  GDRIVE_CREDENTIALS — (optional) service-account JSON for Google Drive
  GDRIVE_FILE_ID    — (optional) spreadsheet file ID on Google Drive
  FEEDBACK_EMAIL    — destination address for member feedback emails
  SMTP_SERVER       — default smtp.gmail.com
  SMTP_PORT         — default 587
  SMTP_USER         — sending email address
  SMTP_PASSWORD     — SMTP / app-specific password
"""

import io
import os
import shutil
import sys
import tempfile
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path

import json

import openpyxl
import pandas as pd
import requests
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Forus Toolkit",
    page_icon="🛡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Language state (must initialise before anything calls t()) ────────────────
st.session_state.setdefault("lang", "EN")

# ── Translations ──────────────────────────────────────────────────────────────
TRANSLATIONS = {
    "EN": {
        # Sidebar
        "nav_dashboard":          "📊 Dashboard",
        "nav_check_mechanisms":   "🔍 Check Annexes",
        "nav_review_queue":       "📋 Review Queue",
        "nav_apply_approved":     "✅ Apply Approved",
        "nav_generate_pdf":       "📄 Generate PDF",
        "nav_give_feedback":      "💬 Give Feedback",
        "nav_manage_tools":       "🔧 Manage Tools",
        "sidebar_download_sp":    "⬇ Download updated spreadsheet",
        "sidebar_save_gdrive":    "☁ Save to Google Drive",
        "sidebar_no_sp":          "No spreadsheet loaded",
        "sidebar_upload_sp":      "Upload spreadsheet (.xlsx)",
        "sidebar_recent_actions": "Recent actions",
        "sidebar_logout":         "Sign out",
        "language_label":         "Language",
        # Login
        "login_title":       "🛡 Forus Toolkit",
        "login_subtitle":    "Enter your password to continue.",
        "login_btn":         "Sign in",
        "login_error":       "Incorrect password — please try again.",
        # Dashboard
        "page_dashboard":        "Forus Toolkit — Dashboard",
        "dash_content_blocks":   "Content blocks",
        "dash_over_limit":       "Over word limit",
        "dash_awaiting_review":  "Awaiting review",
        "dash_mechanisms_due":   "Annexes due",
        "dash_mechanisms":       "Annexes",
        "dash_review_queue":     "Review Queue",
        # Check Mechanisms
        "page_check_mechanisms": "Check Annexes",
        "check_desc": (
            "Run AI-powered verification against each mechanism entry whose "
            "next check date is within the next 14 days, or is flagged **VERIFY**. "
            "Results are written to the **REVIEW_QUEUE** tab of the spreadsheet."
        ),
        "check_no_due":           "✓ No annex entries are currently due for verification.",
        "check_due_label":        "mechanism(s) due for verification:",
        "check_run_verification": "Run verification",
        "check_api_key":          "Anthropic API key",
        "check_run_btn":          "🔍 Run checks now",
        "check_done_msg":         "Done — {no_change} unchanged, {changed} change(s) detected, {unable} unable to verify.",
        "check_review_hint":      "Switch to the **Review Queue** page to review and approve proposed changes.",
        # Review Queue
        "page_review_queue":  "Review Queue",
        "rq_filter_status":   "Filter by status",
        "rq_filter_cat":      "Filter by category",
        "rq_no_items":        "No items match the current filter.",
        "rq_items_shown":     "item(s) shown",
        "rq_current_value":   "Current value",
        "rq_proposed_value":  "Proposed value",
        "rq_approve":         "✅ Approve",
        "rq_reject":          "❌ Reject",
        "rq_saved_hint":      "✓ Saved. Use **Download updated spreadsheet** in the sidebar to keep your changes.",
        # Apply Approved
        "page_apply_approved": "Apply Approved Changes",
        "apply_desc": (
            "Write all **APPROVED** items from the Review Queue back to the "
            "ANNEXES sheet, and update verification dates."
        ),
        "apply_no_approved":   "No APPROVED items in the Review Queue. Approve items on the **Review Queue** page first.",
        "apply_ready":         "{n} approved item(s) ready to apply.",
        "apply_reviewer_name": "Your name (recorded in spreadsheet)",
        "apply_btn":           "✅ Apply all approved changes",
        "apply_done":          "✓ {n} change(s) applied to ANNEXES sheet. Verification dates updated. Review Queue items marked COMPLETED.",
        "apply_download_hint": "Use **Download updated spreadsheet** in the sidebar to save your changes, then regenerate the PDF when ready.",
        # Generate PDF
        "page_generate_pdf":     "Generate PDF",
        "pdf_tab_standard":      "Standard build",
        "pdf_tab_custom":        "Custom member PDF",
        "pdf_standard_desc":     "Generates both the **Public** and **Network** versions of the full toolkit PDF.",
        "pdf_build_btn":         "📄 Build PDF(s)",
        "pdf_download_public":   "⬇ Download Public PDF",
        "pdf_download_network":  "⬇ Download Network PDF",
        "pdf_custom_desc": (
            "Fill in the form below to generate a personalised PDF for a specific member. "
            "Select only the sections relevant to them."
        ),
        "pdf_contact_name":      "Contact name *",
        "pdf_organisation":      "Organisation *",
        "pdf_email":             "Email (optional)",
        "pdf_access_level":      "Access level",
        "pdf_parts_include":     "**Parts to include**",
        "pdf_annexes":           "**Annexes**",
        "pdf_regions":           "**Regions** *(filters annex content)*",
        "pdf_all_regions":       "All regions",
        "pdf_tools":             "**Tools** *(single-page visual tools appended after main content)*",
        "pdf_appendix_tools":    "**Appendix Tools**",
        "pdf_build_custom_btn":  "📄 Build custom PDF",
        "pdf_no_name_org_error": "Please enter a contact name and organisation.",
        "pdf_gen_failed":        "PDF generation failed. Check that at least one section is selected.",
        # Generate PDF — part / annex / region / tool labels
        "pdf_part_1": "Part 1 — Crisis Guides",
        "pdf_part_2": "Part 2 — Solidarity",
        "pdf_part_3": "Part 3 — Legal Support",
        "pdf_part_4": "Part 4 — Emergency Funding",
        "pdf_part_5": "Part 5 — Safe Comms",
        "pdf_part_6": "Part 6 — Diversification",
        "pdf_part_7": "Part 7 — Feedback",
        "pdf_annex_a": "Annex A: Legal Pro Bono Support",
        "pdf_annex_b": "Annex B: Emergency Grants Mechanisms",
        "pdf_annex_c": "Annex C: Physical & Digital Security Support",
        "pdf_region_africa":  "Africa",
        "pdf_region_asia":    "Asia-Pacific",
        "pdf_region_europe":  "Europe",
        "pdf_region_latam":   "Latin America & Caribbean",
        "pdf_region_pacific": "Pacific",
        "pdf_region_global":  "Global",
        "pdf_tool_t1": "Tool 1 — Compliance Self-Check (B1)",
        "pdf_tool_t2": "Tool 2 — Legal Support Decision Tree (B3)",
        "pdf_tool_t3": "Tool 3 — Go public or stay quiet? (B5)",
        "pdf_tool_t4": "Tool 4 — Do-No-Harm Checklist (B5)",
        "pdf_appendix_a1": "Appendix A1 — Platform Role Clarifier (B2)",
        "pdf_appendix_a2": "Appendix A2 — Diversification Readiness Gate (B6)",
        "pdf_appendix_a3": "Appendix A3 — Emergency Funding Navigator (B4)",
        "pdf_access_public_opt":  "Public",
        "pdf_access_network_opt": "Network",
        # Give Feedback
        "page_give_feedback":     "Give Feedback",
        "feedback_member_desc": (
            "Use this form to flag anything in the toolkit that seems outdated, "
            "incorrect, or missing. Your message goes directly to the Forus Secretariat."
        ),
        "feedback_name_label":    "Your name (optional)",
        "feedback_org_label":     "Your organisation (optional)",
        "feedback_part_label":    "Which part of the toolkit?",
        "feedback_section_label": "Which section (optional — be as specific as you like)",
        "feedback_issue_label":   "Type of issue",
        "feedback_issue_opt_1":   "Outdated information",
        "feedback_issue_opt_2":   "Missing information",
        "feedback_issue_opt_3":   "Factual error",
        "feedback_issue_opt_4":   "Broken or missing link",
        "feedback_issue_opt_5":   "Other",
        "feedback_desc_label":    "Please describe the issue *",
        "feedback_submit_btn":    "📨 Send feedback",
        "feedback_success":       "✓ Thank you — your feedback has been sent to the Forus Secretariat.",
        "feedback_empty_error":   "Please describe the issue before submitting.",
        "feedback_send_error":    "Could not send email: {err}. Please contact the Secretariat directly.",
        "feedback_part_general":  "General (not section-specific)",
        "feedback_admin_tab_submit":  "Submit feedback",
        "feedback_admin_tab_view":    "View submitted feedback",
        "feedback_admin_no_entries":  "No member feedback has been submitted yet.",
        "feedback_admin_col_time":    "Submitted",
        "feedback_admin_col_name":    "Name",
        "feedback_admin_col_org":     "Organisation",
        "feedback_admin_col_part":    "Part",
        "feedback_admin_col_issue":   "Issue type",
        "feedback_admin_col_desc":    "Description",
        "feedback_admin_col_status":  "Status",
        # Manage Tools
        "page_manage_tools": "Manage Tools",
        "tools_desc": (
            "View and edit the text content for the seven visual tools (T1–T4, A1–A3). "
            "Content is loaded from the **TOOLS** sheet of your spreadsheet. "
            "Changes saved here will appear in the next PDF you generate."
        ),
        "tools_filter":        "Filter by tool",
        "tools_show_flagged":  "Show only flagged entries (VERIFY / UPDATED)",
        "tools_fields_shown":  "field(s) shown",
        "tools_content_label": "Content",
        "tools_status":        "Status",
        "tools_save_btn":      "💾 Save this field",
        "tools_instructions": (
            "**How to use the AI auto-update agent with tools:**\n\n"
            "1. Set any field's status to **VERIFY** if you think the content may be out of date.\n"
            "2. The Check Mechanisms agent reviews ANNEXES entries automatically.\n"
            "   For tool text, flag fields manually and edit them directly here, "
            "   or ask the AI agent in a conversation to review specific fields.\n"
            "3. After editing, use **Save to Google Drive** in the sidebar to persist changes."
        ),
        # General
        "no_spreadsheet_warning": "Please upload the spreadsheet using the sidebar to get started.",
    },

    "FR": {
        # Sidebar
        "nav_dashboard":          "📊 Tableau de bord",
        "nav_check_mechanisms":   "🔍 Vérifier annexes",
        "nav_review_queue":       "📋 File de révision",
        "nav_apply_approved":     "✅ Appliquer approuvés",
        "nav_generate_pdf":       "📄 Générer PDF",
        "nav_give_feedback":      "💬 Donner un avis",
        "nav_manage_tools":       "🔧 Gérer les outils",
        "sidebar_download_sp":    "⬇ Télécharger la feuille de calcul",
        "sidebar_save_gdrive":    "☁ Sauvegarder sur Google Drive",
        "sidebar_no_sp":          "Aucune feuille de calcul chargée",
        "sidebar_upload_sp":      "Importer la feuille de calcul (.xlsx)",
        "sidebar_recent_actions": "Actions récentes",
        "sidebar_logout":         "Se déconnecter",
        "language_label":         "Langue",
        # Login
        "login_title":    "🛡 Boîte à Outils Forus",
        "login_subtitle": "Entrez votre mot de passe pour continuer.",
        "login_btn":      "Se connecter",
        "login_error":    "Mot de passe incorrect — veuillez réessayer.",
        # Dashboard
        "page_dashboard":        "Boîte à Outils Forus — Tableau de bord",
        "dash_content_blocks":   "Blocs de contenu",
        "dash_over_limit":       "Au-dessus de la limite",
        "dash_awaiting_review":  "En attente de révision",
        "dash_mechanisms_due":   "Annexes à vérifier",
        "dash_mechanisms":       "Annexes",
        "dash_review_queue":     "File de révision",
        # Check Mechanisms
        "page_check_mechanisms": "Vérifier les annexes",
        "check_desc": (
            "Lancez une vérification assistée par IA pour chaque mécanisme dont la prochaine "
            "date de vérification est dans les 14 prochains jours, ou marqué **VERIFY**. "
            "Les résultats sont écrits dans l'onglet **REVIEW_QUEUE** de la feuille de calcul."
        ),
        "check_no_due":           "✓ Aucun mécanisme n'est actuellement à vérifier.",
        "check_due_label":        "mécanisme(s) à vérifier :",
        "check_run_verification": "Lancer la vérification",
        "check_api_key":          "Clé API Anthropic",
        "check_run_btn":          "🔍 Lancer les vérifications",
        "check_done_msg":         "Terminé — {no_change} inchangé(s), {changed} modification(s) détectée(s), {unable} impossible(s) à vérifier.",
        "check_review_hint":      "Allez sur la page **File de révision** pour examiner et approuver les modifications proposées.",
        # Review Queue
        "page_review_queue":  "File de révision",
        "rq_filter_status":   "Filtrer par statut",
        "rq_filter_cat":      "Filtrer par catégorie",
        "rq_no_items":        "Aucun élément ne correspond au filtre actuel.",
        "rq_items_shown":     "élément(s) affiché(s)",
        "rq_current_value":   "Valeur actuelle",
        "rq_proposed_value":  "Valeur proposée",
        "rq_approve":         "✅ Approuver",
        "rq_reject":          "❌ Rejeter",
        "rq_saved_hint":      "✓ Sauvegardé. Utilisez **Télécharger la feuille de calcul** dans la barre latérale pour conserver vos modifications.",
        # Apply Approved
        "page_apply_approved": "Appliquer les modifications approuvées",
        "apply_desc": (
            "Écrit tous les éléments **APPROUVÉS** de la file de révision dans "
            "la feuille ANNEXES et met à jour les dates de vérification."
        ),
        "apply_no_approved":   "Aucun élément APPROUVÉ dans la file. Approuvez des éléments sur la page **File de révision**.",
        "apply_ready":         "{n} élément(s) approuvé(s) prêt(s) à appliquer.",
        "apply_reviewer_name": "Votre nom (enregistré dans la feuille)",
        "apply_btn":           "✅ Appliquer toutes les modifications approuvées",
        "apply_done":          "✓ {n} modification(s) appliquée(s). Dates de vérification mises à jour. Éléments marqués COMPLETED.",
        "apply_download_hint": "Utilisez **Télécharger la feuille de calcul** pour sauvegarder vos modifications, puis régénérez le PDF.",
        # Generate PDF
        "page_generate_pdf":    "Générer le PDF",
        "pdf_tab_standard":     "Génération standard",
        "pdf_tab_custom":       "PDF personnalisé pour membre",
        "pdf_standard_desc":    "Génère les versions **Publique** et **Réseau** du PDF complet de la boîte à outils.",
        "pdf_build_btn":        "📄 Générer le(s) PDF",
        "pdf_download_public":  "⬇ Télécharger le PDF Public",
        "pdf_download_network": "⬇ Télécharger le PDF Réseau",
        "pdf_custom_desc": (
            "Remplissez le formulaire ci-dessous pour générer un PDF personnalisé pour un membre spécifique. "
            "Sélectionnez uniquement les sections qui lui sont pertinentes."
        ),
        "pdf_contact_name":      "Nom du contact *",
        "pdf_organisation":      "Organisation *",
        "pdf_email":             "E-mail (facultatif)",
        "pdf_access_level":      "Niveau d'accès",
        "pdf_parts_include":     "**Parties à inclure**",
        "pdf_annexes":           "**Annexes**",
        "pdf_regions":           "**Régions** *(filtre le contenu des annexes)*",
        "pdf_all_regions":       "Toutes les régions",
        "pdf_tools":             "**Outils** *(outils visuels d'une page ajoutés après le contenu principal)*",
        "pdf_appendix_tools":    "**Outils en annexe**",
        "pdf_build_custom_btn":  "📄 Générer le PDF personnalisé",
        "pdf_no_name_org_error": "Veuillez entrer un nom de contact et une organisation.",
        "pdf_gen_failed":        "La génération du PDF a échoué. Vérifiez qu'au moins une section est sélectionnée.",
        "pdf_part_1": "Partie 1 — Guides de crise",
        "pdf_part_2": "Partie 2 — Solidarité",
        "pdf_part_3": "Partie 3 — Soutien juridique",
        "pdf_part_4": "Partie 4 — Financement d'urgence",
        "pdf_part_5": "Partie 5 — Comms sécurisées",
        "pdf_part_6": "Partie 6 — Diversification",
        "pdf_part_7": "Partie 7 — Retours",
        "pdf_annex_a": "Annexe A : Soutien bénévole juridique",
        "pdf_annex_b": "Annexe B : Mécanismes de subventions d'urgence",
        "pdf_annex_c": "Annexe C : Soutien à la sécurité physique et numérique",
        "pdf_region_africa":  "Afrique",
        "pdf_region_asia":    "Asie-Pacifique",
        "pdf_region_europe":  "Europe",
        "pdf_region_latam":   "Amérique latine et Caraïbes",
        "pdf_region_pacific": "Pacifique",
        "pdf_region_global":  "Mondial",
        "pdf_tool_t1": "Outil 1 — Auto-vérification de conformité (B1)",
        "pdf_tool_t2": "Outil 2 — Arbre de décision soutien juridique (B3)",
        "pdf_tool_t3": "Outil 3 — Rendre public ou rester discret ? (B5)",
        "pdf_tool_t4": "Outil 4 — Liste de contrôle Do-No-Harm (B5)",
        "pdf_appendix_a1": "Annexe A1 — Clarificateur de rôle de plateforme (B2)",
        "pdf_appendix_a2": "Annexe A2 — Passerelle de préparation à la diversification (B6)",
        "pdf_appendix_a3": "Annexe A3 — Navigateur de financement d'urgence (B4)",
        "pdf_access_public_opt":  "Public",
        "pdf_access_network_opt": "Réseau",
        # Give Feedback
        "page_give_feedback":    "Donner un avis",
        "feedback_member_desc": (
            "Utilisez ce formulaire pour signaler tout élément de la boîte à outils qui semble "
            "obsolète, incorrect ou manquant. Votre message est transmis directement au Secrétariat de Forus."
        ),
        "feedback_name_label":    "Votre nom (facultatif)",
        "feedback_org_label":     "Votre organisation (facultatif)",
        "feedback_part_label":    "Quelle partie de la boîte à outils ?",
        "feedback_section_label": "Quelle section (facultatif — soyez aussi précis que vous le souhaitez)",
        "feedback_issue_label":   "Type de problème",
        "feedback_issue_opt_1":   "Information obsolète",
        "feedback_issue_opt_2":   "Information manquante",
        "feedback_issue_opt_3":   "Erreur factuelle",
        "feedback_issue_opt_4":   "Lien brisé ou manquant",
        "feedback_issue_opt_5":   "Autre",
        "feedback_desc_label":    "Décrivez le problème *",
        "feedback_submit_btn":    "📨 Envoyer le retour",
        "feedback_success":       "✓ Merci — votre retour a été envoyé au Secrétariat de Forus.",
        "feedback_empty_error":   "Veuillez décrire le problème avant de soumettre.",
        "feedback_send_error":    "Impossible d'envoyer l'e-mail : {err}. Veuillez contacter directement le Secrétariat.",
        "feedback_part_general":  "Général (pas spécifique à une section)",
        "feedback_admin_tab_submit":  "Soumettre un avis",
        "feedback_admin_tab_view":    "Consulter les avis reçus",
        "feedback_admin_no_entries":  "Aucun retour membre n'a encore été soumis.",
        "feedback_admin_col_time":    "Soumis le",
        "feedback_admin_col_name":    "Nom",
        "feedback_admin_col_org":     "Organisation",
        "feedback_admin_col_part":    "Partie",
        "feedback_admin_col_issue":   "Type de problème",
        "feedback_admin_col_desc":    "Description",
        "feedback_admin_col_status":  "Statut",
        # Manage Tools
        "page_manage_tools": "Gérer les outils",
        "tools_desc": (
            "Consultez et modifiez le contenu textuel des sept outils visuels (T1–T4, A1–A3). "
            "Le contenu est chargé depuis l'onglet **TOOLS** de votre feuille de calcul. "
            "Les modifications apportées ici apparaîtront dans le prochain PDF généré."
        ),
        "tools_filter":        "Filtrer par outil",
        "tools_show_flagged":  "Afficher uniquement les entrées signalées (VERIFY / UPDATED)",
        "tools_fields_shown":  "champ(s) affiché(s)",
        "tools_content_label": "Contenu",
        "tools_status":        "Statut",
        "tools_save_btn":      "💾 Sauvegarder ce champ",
        "tools_instructions": (
            "**Comment utiliser l'agent IA de mise à jour avec les outils :**\n\n"
            "1. Définissez le statut d'un champ sur **VERIFY** si vous pensez que le contenu est obsolète.\n"
            "2. L'agent Check Annexes vérifie automatiquement les entrées ANNEXES.\n"
            "   Pour le texte des outils, signalez les champs manuellement et modifiez-les directement ici.\n"
            "3. Après modification, utilisez **Sauvegarder sur Google Drive** pour conserver les changements."
        ),
        "no_spreadsheet_warning": "Veuillez importer la feuille de calcul via la barre latérale pour commencer.",
    },

    "ES": {
        # Sidebar
        "nav_dashboard":          "📊 Panel de control",
        "nav_check_mechanisms":   "🔍 Verificar anexos",
        "nav_review_queue":       "📋 Cola de revisión",
        "nav_apply_approved":     "✅ Aplicar aprobados",
        "nav_generate_pdf":       "📄 Generar PDF",
        "nav_give_feedback":      "💬 Dar retroalimentación",
        "nav_manage_tools":       "🔧 Gestionar herramientas",
        "sidebar_download_sp":    "⬇ Descargar hoja de cálculo",
        "sidebar_save_gdrive":    "☁ Guardar en Google Drive",
        "sidebar_no_sp":          "No se ha cargado ninguna hoja de cálculo",
        "sidebar_upload_sp":      "Subir hoja de cálculo (.xlsx)",
        "sidebar_recent_actions": "Acciones recientes",
        "sidebar_logout":         "Cerrar sesión",
        "language_label":         "Idioma",
        # Login
        "login_title":    "🛡 Kit de Herramientas Forus",
        "login_subtitle": "Ingrese su contraseña para continuar.",
        "login_btn":      "Iniciar sesión",
        "login_error":    "Contraseña incorrecta — por favor, inténtelo de nuevo.",
        # Dashboard
        "page_dashboard":        "Kit de Herramientas Forus — Panel de control",
        "dash_content_blocks":   "Bloques de contenido",
        "dash_over_limit":       "Por encima del límite",
        "dash_awaiting_review":  "Pendiente de revisión",
        "dash_mechanisms_due":   "Anexos pendientes",
        "dash_mechanisms":       "Anexos",
        "dash_review_queue":     "Cola de revisión",
        # Check Mechanisms
        "page_check_mechanisms": "Verificar anexos",
        "check_desc": (
            "Ejecute la verificación asistida por IA para cada mecanismo cuya próxima "
            "fecha de verificación esté dentro de los próximos 14 días o esté marcado como **VERIFY**. "
            "Los resultados se escriben en la pestaña **REVIEW_QUEUE** de la hoja de cálculo."
        ),
        "check_no_due":           "✓ No hay mecanismos pendientes de verificación en este momento.",
        "check_due_label":        "mecanismo(s) pendiente(s) de verificación:",
        "check_run_verification": "Ejecutar verificación",
        "check_api_key":          "Clave API de Anthropic",
        "check_run_btn":          "🔍 Ejecutar verificaciones",
        "check_done_msg":         "Listo — {no_change} sin cambios, {changed} cambio(s) detectado(s), {unable} no verificable(s).",
        "check_review_hint":      "Vaya a la página **Cola de revisión** para revisar y aprobar los cambios propuestos.",
        # Review Queue
        "page_review_queue":  "Cola de revisión",
        "rq_filter_status":   "Filtrar por estado",
        "rq_filter_cat":      "Filtrar por categoría",
        "rq_no_items":        "Ningún elemento coincide con el filtro actual.",
        "rq_items_shown":     "elemento(s) mostrado(s)",
        "rq_current_value":   "Valor actual",
        "rq_proposed_value":  "Valor propuesto",
        "rq_approve":         "✅ Aprobar",
        "rq_reject":          "❌ Rechazar",
        "rq_saved_hint":      "✓ Guardado. Use **Descargar hoja de cálculo** en la barra lateral para conservar sus cambios.",
        # Apply Approved
        "page_apply_approved": "Aplicar cambios aprobados",
        "apply_desc": (
            "Escribe todos los elementos **APROBADOS** de la cola de revisión en la "
            "hoja ANNEXES y actualiza las fechas de verificación."
        ),
        "apply_no_approved":   "No hay elementos APROBADOS. Apruebe elementos en la página **Cola de revisión** primero.",
        "apply_ready":         "{n} elemento(s) aprobado(s) listo(s) para aplicar.",
        "apply_reviewer_name": "Su nombre (registrado en la hoja de cálculo)",
        "apply_btn":           "✅ Aplicar todos los cambios aprobados",
        "apply_done":          "✓ {n} cambio(s) aplicado(s) a la hoja ANNEXES. Fechas actualizadas. Elementos marcados como COMPLETED.",
        "apply_download_hint": "Use **Descargar hoja de cálculo** para guardar sus cambios y luego regenere el PDF.",
        # Generate PDF
        "page_generate_pdf":    "Generar PDF",
        "pdf_tab_standard":     "Generación estándar",
        "pdf_tab_custom":       "PDF personalizado para miembro",
        "pdf_standard_desc":    "Genera las versiones **Pública** y **de Red** del PDF completo del kit de herramientas.",
        "pdf_build_btn":        "📄 Generar PDF(s)",
        "pdf_download_public":  "⬇ Descargar PDF Público",
        "pdf_download_network": "⬇ Descargar PDF de Red",
        "pdf_custom_desc": (
            "Complete el formulario a continuación para generar un PDF personalizado para un miembro específico. "
            "Seleccione solo las secciones relevantes para ese miembro."
        ),
        "pdf_contact_name":      "Nombre del contacto *",
        "pdf_organisation":      "Organización *",
        "pdf_email":             "Correo electrónico (opcional)",
        "pdf_access_level":      "Nivel de acceso",
        "pdf_parts_include":     "**Partes a incluir**",
        "pdf_annexes":           "**Anexos**",
        "pdf_regions":           "**Regiones** *(filtra el contenido de los anexos)*",
        "pdf_all_regions":       "Todas las regiones",
        "pdf_tools":             "**Herramientas** *(herramientas visuales de una página añadidas tras el contenido principal)*",
        "pdf_appendix_tools":    "**Herramientas de apéndice**",
        "pdf_build_custom_btn":  "📄 Generar PDF personalizado",
        "pdf_no_name_org_error": "Por favor, ingrese un nombre de contacto y una organización.",
        "pdf_gen_failed":        "La generación del PDF falló. Compruebe que haya al menos una sección seleccionada.",
        "pdf_part_1": "Parte 1 — Guías de crisis",
        "pdf_part_2": "Parte 2 — Solidaridad",
        "pdf_part_3": "Parte 3 — Apoyo jurídico",
        "pdf_part_4": "Parte 4 — Financiación de emergencia",
        "pdf_part_5": "Parte 5 — Comunicaciones seguras",
        "pdf_part_6": "Parte 6 — Diversificación",
        "pdf_part_7": "Parte 7 — Retroalimentación",
        "pdf_annex_a": "Anexo A: Apoyo jurídico pro bono",
        "pdf_annex_b": "Anexo B: Mecanismos de subvenciones de emergencia",
        "pdf_annex_c": "Anexo C: Apoyo a la seguridad física y digital",
        "pdf_region_africa":  "África",
        "pdf_region_asia":    "Asia-Pacífico",
        "pdf_region_europe":  "Europa",
        "pdf_region_latam":   "América Latina y el Caribe",
        "pdf_region_pacific": "Pacífico",
        "pdf_region_global":  "Global",
        "pdf_tool_t1": "Herramienta 1 — Autocomprobación de cumplimiento (B1)",
        "pdf_tool_t2": "Herramienta 2 — Árbol de decisión de apoyo jurídico (B3)",
        "pdf_tool_t3": "Herramienta 3 — ¿Hacerlo público o mantenerlo en silencio? (B5)",
        "pdf_tool_t4": "Herramienta 4 — Lista de verificación Do-No-Harm (B5)",
        "pdf_appendix_a1": "Apéndice A1 — Clarificador de rol de plataforma (B2)",
        "pdf_appendix_a2": "Apéndice A2 — Puerta de preparación para la diversificación (B6)",
        "pdf_appendix_a3": "Apéndice A3 — Navegador de financiación de emergencia (B4)",
        "pdf_access_public_opt":  "Público",
        "pdf_access_network_opt": "Red",
        # Give Feedback
        "page_give_feedback":    "Dar retroalimentación",
        "feedback_member_desc": (
            "Use este formulario para señalar cualquier elemento del kit de herramientas que parezca "
            "desactualizado, incorrecto o que falte. Su mensaje se envía directamente al Secretariado de Forus."
        ),
        "feedback_name_label":    "Su nombre (opcional)",
        "feedback_org_label":     "Su organización (opcional)",
        "feedback_part_label":    "¿Qué parte del kit de herramientas?",
        "feedback_section_label": "¿Qué sección? (opcional — sea tan específico como desee)",
        "feedback_issue_label":   "Tipo de problema",
        "feedback_issue_opt_1":   "Información desactualizada",
        "feedback_issue_opt_2":   "Información faltante",
        "feedback_issue_opt_3":   "Error factual",
        "feedback_issue_opt_4":   "Enlace roto o faltante",
        "feedback_issue_opt_5":   "Otro",
        "feedback_desc_label":    "Describa el problema *",
        "feedback_submit_btn":    "📨 Enviar retroalimentación",
        "feedback_success":       "✓ Gracias — sus comentarios han sido enviados al Secretariado de Forus.",
        "feedback_empty_error":   "Por favor, describa el problema antes de enviar.",
        "feedback_send_error":    "No se pudo enviar el correo: {err}. Contacte directamente al Secretariado.",
        "feedback_part_general":  "General (no específico de una sección)",
        "feedback_admin_tab_submit":  "Enviar comentario",
        "feedback_admin_tab_view":    "Ver retroalimentación recibida",
        "feedback_admin_no_entries":  "Aún no se ha enviado ninguna retroalimentación de miembros.",
        "feedback_admin_col_time":    "Enviado",
        "feedback_admin_col_name":    "Nombre",
        "feedback_admin_col_org":     "Organización",
        "feedback_admin_col_part":    "Parte",
        "feedback_admin_col_issue":   "Tipo de problema",
        "feedback_admin_col_desc":    "Descripción",
        "feedback_admin_col_status":  "Estado",
        # Manage Tools
        "page_manage_tools": "Gestionar herramientas",
        "tools_desc": (
            "Vea y edite el contenido textual de las siete herramientas visuales (T1–T4, A1–A3). "
            "El contenido se carga desde la pestaña **TOOLS** de su hoja de cálculo. "
            "Los cambios realizados aquí aparecerán en el próximo PDF generado."
        ),
        "tools_filter":        "Filtrar por herramienta",
        "tools_show_flagged":  "Mostrar solo entradas marcadas (VERIFY / UPDATED)",
        "tools_fields_shown":  "campo(s) mostrado(s)",
        "tools_content_label": "Contenido",
        "tools_status":        "Estado",
        "tools_save_btn":      "💾 Guardar este campo",
        "tools_instructions": (
            "**Cómo usar el agente de IA de actualización automática con las herramientas:**\n\n"
            "1. Establezca el estado de un campo en **VERIFY** si cree que el contenido puede estar desactualizado.\n"
            "2. El agente Check Annexes revisa las entradas de ANNEXES automáticamente.\n"
            "   Para el texto de las herramientas, marque los campos manualmente y edítelos directamente aquí.\n"
            "3. Después de editar, use **Guardar en Google Drive** para conservar los cambios."
        ),
        "no_spreadsheet_warning": "Por favor, suba la hoja de cálculo usando la barra lateral para comenzar.",
    },
}


def t(key):
    """Return translated string for the current UI language."""
    lang = st.session_state.get("lang", "EN")
    return (
        TRANSLATIONS.get(lang, TRANSLATIONS["EN"]).get(key)
        or TRANSLATIONS["EN"].get(key, key)
    )


# ── Role constants ────────────────────────────────────────────────────────────
ROLE_ADMIN  = "admin"
ROLE_MEMBER = "member"

# Nav keys visible to each role
_ADMIN_NAV_KEYS  = [
    "dashboard", "check_mechanisms", "review_queue",
    "apply_approved", "generate_pdf", "give_feedback", "manage_tools",
]
_MEMBER_NAV_KEYS = ["generate_pdf", "give_feedback"]


# ── Password / role gate ───────────────────────────────────────────────────────
def _check_password():
    """Show login screen; set role in session state on success."""
    if st.session_state.get("authenticated"):
        return
    st.markdown(
        f"""
        <div style='max-width:400px;margin:80px auto 0;text-align:center'>
            <h2 style='color:#00424D'>{t("login_title")}</h2>
            <p style='color:#555;margin-bottom:1.5rem'>{t("login_subtitle")}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pwd = st.text_input("Password", type="password",
                            label_visibility="collapsed", placeholder="Password")
        if st.button(t("login_btn"), use_container_width=True, type="primary"):
            admin_pwd  = st.secrets.get("APP_PASSWORD", "")
            member_pwd = st.secrets.get("MEMBER_PASSWORD", "")
            if admin_pwd and pwd == admin_pwd:
                st.session_state.authenticated = True
                st.session_state.role = ROLE_ADMIN
                st.rerun()
            elif member_pwd and pwd == member_pwd:
                st.session_state.authenticated = True
                st.session_state.role = ROLE_MEMBER
                st.rerun()
            else:
                st.error(t("login_error"))
    st.stop()

_check_password()

# ── Brand colours ─────────────────────────────────────────────────────────────
TEAL  = "#58C5C7"
DARK  = "#00424D"
PINK  = "#ED1651"
MINT  = "#5C9C8E"
LIME  = "#B2C100"
LGREY = "#F5F5F5"

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  [data-testid="stSidebar"] {{
    background-color: {DARK};
  }}
  [data-testid="stSidebar"] * {{
    color: white !important;
  }}
  [data-testid="stSidebar"] .stRadio label {{
    font-size: 15px;
    padding: 4px 0;
  }}
  /* Sidebar buttons: visible on dark background */
  [data-testid="stSidebar"] button {{
    background-color: rgba(255,255,255,0.15) !important;
    color: white !important;
    border: 1px solid rgba(255,255,255,0.35) !important;
  }}
  [data-testid="stSidebar"] button:hover {{
    background-color: rgba(255,255,255,0.25) !important;
  }}
  /* Sidebar selectbox */
  [data-testid="stSidebar"] [data-baseweb="select"] > div:first-child {{
    background-color: rgba(255,255,255,0.15) !important;
    border-color: rgba(255,255,255,0.35) !important;
  }}
  [data-testid="stSidebar"] [data-baseweb="select"] svg {{
    fill: white !important;
  }}
  [data-testid="stSidebar"] [data-testid="stFileUploader"] label {{
    color: white !important;
  }}
  /* Page headings */
  h1 {{ color: {DARK}; border-bottom: 3px solid {TEAL}; padding-bottom: 6px; }}
  h2 {{ color: {DARK}; }}
  h3 {{ color: {MINT}; }}
  /* Cards */
  .forus-card {{
    background: {LGREY};
    border-left: 4px solid {TEAL};
    border-radius: 4px;
    padding: 14px 16px;
    margin-bottom: 12px;
  }}
  .forus-card.change   {{ border-color: {PINK}; }}
  .forus-card.approved {{ border-color: {MINT}; background: #E8F5F2; }}
  .forus-card.rejected {{ border-color: #888; background: #F9F9F9; opacity: 0.7; }}
  /* Confidence badges */
  .badge {{
    display: inline-block;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: bold;
    color: white;
  }}
  .badge.HIGH   {{ background: {MINT}; }}
  .badge.MEDIUM {{ background: {LIME}; color: {DARK}; }}
  .badge.LOW    {{ background: {PINK}; }}
  /* Stat boxes */
  .stat-box {{
    background: white;
    border: 1px solid #DDD;
    border-top: 4px solid {TEAL};
    border-radius: 4px;
    padding: 16px;
    text-align: center;
  }}
  .stat-num {{ font-size: 32px; font-weight: bold; color: {DARK}; }}
  .stat-lbl {{ font-size: 13px; color: #666; margin-top: 4px; }}
</style>
""", unsafe_allow_html=True)


# ── Session state ─────────────────────────────────────────────────────────────
def _init():
    role = st.session_state.get("role", ROLE_ADMIN)
    if "sp_path" not in st.session_state:
        # Members always load from Google Drive; admins can also use local file
        if role == ROLE_MEMBER:
            st.session_state["sp_path"] = None
            st.session_state["sp_name"] = None
        else:
            local = Path("Forus_Toolkit_Content_DB.xlsx")
            if local.exists() and not st.secrets.get("GDRIVE_CREDENTIALS"):
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                shutil.copy(str(local), tmp.name)
                tmp.close()
                st.session_state["sp_path"] = tmp.name
                st.session_state["sp_name"] = local.name
            else:
                st.session_state["sp_path"] = None
                st.session_state["sp_name"] = None
    if "action_log" not in st.session_state:
        st.session_state["action_log"] = []

_init()


# ── Helpers ───────────────────────────────────────────────────────────────────
def sp():
    return st.session_state.get("sp_path")

def is_admin():
    return st.session_state.get("role") == ROLE_ADMIN

def _load_wb(data_only=True):
    return openpyxl.load_workbook(sp(), data_only=data_only)

def _save_wb(wb):
    wb.save(sp())

def _sheet_to_df(wb, sheet_name, header_row=2):
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws   = wb[sheet_name]
    hdrs = [c.value for c in ws[header_row]]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if any(v is not None for v in row):
            rows.append({hdrs[i]: row[i] for i in range(len(hdrs)) if i < len(row)})
    return pd.DataFrame(rows)

def _badge(conf):
    conf = str(conf or "").upper()
    return f'<span class="badge {conf}">{conf}</span>'


# ── Google Drive helpers ──────────────────────────────────────────────────────
def _gdrive_service():
    creds_json = st.secrets.get("GDRIVE_CREDENTIALS", "")
    if not creds_json:
        return None
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        creds_dict = json.loads(creds_json)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict, scopes=["https://www.googleapis.com/auth/drive"],
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        st.warning(f"Could not build Drive service: {e}")
        return None


def _load_from_gdrive():
    if st.session_state.get("sp_path"):
        return
    file_id = st.secrets.get("GDRIVE_FILE_ID", "")
    if not file_id:
        return
    svc  = _gdrive_service()
    data = None
    if svc:
        try:
            from googleapiclient.http import MediaIoBaseDownload
            meta = svc.files().get(fileId=file_id, fields="mimeType").execute()
            mime = meta.get("mimeType", "")
            if mime == "application/vnd.google-apps.spreadsheet":
                request = svc.files().export_media(
                    fileId=file_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                request = svc.files().get_media(fileId=file_id)
            buf = io.BytesIO()
            dl  = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = dl.next_chunk()
            data = buf.getvalue()
        except Exception as sa_err:
            st.warning(f"Service account access failed ({sa_err}). Trying link-share fallback...")
    if data is None:
        try:
            url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
            r   = requests.get(url, timeout=30, allow_redirects=True)
            if r.status_code == 200 and len(r.content) > 1000:
                data = r.content
            else:
                url = f"https://drive.google.com/uc?export=download&id={file_id}"
                r   = requests.get(url, timeout=30, allow_redirects=True)
                r.raise_for_status()
                data = r.content
        except Exception as pub_err:
            st.warning(f"Could not load spreadsheet from Google Drive: {pub_err}")
            return
    if data:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(data)
        tmp.close()
        st.session_state["sp_path"] = tmp.name
        st.session_state["sp_name"] = "Forus_Toolkit_Content_DB.xlsx (Google Drive)"


def save_to_gdrive():
    file_id = st.secrets.get("GDRIVE_FILE_ID", "")
    if not file_id or not sp():
        return False, "No file ID or spreadsheet path configured."
    svc = _gdrive_service()
    if not svc:
        return False, "No service-account credentials found."
    try:
        from googleapiclient.http import MediaFileUpload
        media = MediaFileUpload(
            sp(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False,
        )
        svc.files().update(
            fileId=file_id, media_body=media, supportsAllDrives=True,
        ).execute()
        return True, "Saved to Google Drive."
    except Exception as e:
        return False, str(e)


_load_from_gdrive()


# ── Member feedback helpers ───────────────────────────────────────────────────
_MF_SHEET   = "MEMBER_FEEDBACK"
_MF_COLS    = ["timestamp", "name", "organisation", "language",
               "part", "section", "issue_type", "description", "status"]

def _ensure_mf_sheet(wb):
    if _MF_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_MF_SHEET)
        ws.cell(1, 1).value = "Member Feedback"
        for i, h in enumerate(_MF_COLS, 1):
            ws.cell(2, i).value = h
    return wb[_MF_SHEET]


def _save_feedback_to_spreadsheet(data: dict) -> bool:
    """Append a feedback row to the MEMBER_FEEDBACK sheet and save."""
    if not sp():
        return False
    try:
        wb = _load_wb(data_only=False)
        ws = _ensure_mf_sheet(wb)
        ws.append([data.get(c, "") for c in _MF_COLS])
        _save_wb(wb)
        if st.secrets.get("GDRIVE_CREDENTIALS"):
            save_to_gdrive()
        return True
    except Exception:
        return False


def _send_feedback_email(data: dict) -> tuple:
    """Email feedback to the Secretariat. Returns (success, message)."""
    to_addr     = st.secrets.get("FEEDBACK_EMAIL", "")
    smtp_server = st.secrets.get("SMTP_SERVER", "smtp.gmail.com")
    smtp_port   = int(st.secrets.get("SMTP_PORT", 587))
    smtp_user   = st.secrets.get("SMTP_USER", "")
    smtp_pass   = st.secrets.get("SMTP_PASSWORD", "")

    if not to_addr or not smtp_user or not smtp_pass:
        return False, (
            "Email not configured — add FEEDBACK_EMAIL, SMTP_USER and "
            "SMTP_PASSWORD to Streamlit secrets."
        )

    subject = f"[Forus Toolkit] Feedback: {data.get('issue_type', 'Update needed')} — {data.get('part', '')}"
    body = (
        f"Forus Toolkit Member Feedback\n"
        f"{'=' * 40}\n\n"
        f"Submitted:    {data.get('timestamp', '')}\n"
        f"Name:         {data.get('name', 'Anonymous') or 'Anonymous'}\n"
        f"Organisation: {data.get('organisation', 'Not provided') or 'Not provided'}\n"
        f"Language:     {data.get('language', 'EN')}\n\n"
        f"Part:         {data.get('part', '')}\n"
        f"Section:      {data.get('section', 'Not specified') or 'Not specified'}\n"
        f"Issue type:   {data.get('issue_type', '')}\n\n"
        f"Description:\n{data.get('description', '')}\n"
    )

    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return True, "Email sent."
    except Exception as e:
        return False, str(e)


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🛡 Forus Toolkit")

    # Language picker
    _lang_options = {"🇬🇧 English": "EN", "🇫🇷 Français": "FR", "🇪🇸 Español": "ES"}
    _lang_display = list(_lang_options.keys())
    _current_lang_display = next(
        (k for k, v in _lang_options.items() if v == st.session_state.get("lang", "EN")),
        "🇬🇧 English",
    )
    _selected_lang_display = st.selectbox(
        t("language_label"), _lang_display,
        index=_lang_display.index(_current_lang_display),
        label_visibility="collapsed",
    )
    _new_lang = _lang_options[_selected_lang_display]
    if _new_lang != st.session_state.get("lang"):
        st.session_state["lang"] = _new_lang
        st.rerun()

    st.markdown("---")

    # Spreadsheet controls — admin only
    if is_admin():
        if sp():
            st.success(f"📊 {st.session_state['sp_name']}")
            with open(sp(), "rb") as f:
                st.download_button(
                    t("sidebar_download_sp"), f.read(),
                    file_name="Forus_Toolkit_Content_DB.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            if st.secrets.get("GDRIVE_CREDENTIALS"):
                if st.button(t("sidebar_save_gdrive"), use_container_width=True):
                    ok, msg = save_to_gdrive()
                    (st.success if ok else st.error)(msg)
        else:
            st.warning(t("sidebar_no_sp"))
            uploaded = st.file_uploader(t("sidebar_upload_sp"), type=["xlsx"])
            if uploaded:
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp.write(uploaded.read())
                tmp.close()
                st.session_state["sp_path"] = tmp.name
                st.session_state["sp_name"] = uploaded.name
                st.rerun()
        st.markdown("---")

    # Navigation — filtered by role
    _nav_keys    = _ADMIN_NAV_KEYS if is_admin() else _MEMBER_NAV_KEYS
    _nav_display = [t(f"nav_{k}") for k in _nav_keys]
    _selected    = st.radio("Navigation", _nav_display, label_visibility="collapsed")
    page         = _nav_keys[_nav_display.index(_selected)]

    # Recent actions (admin only)
    if is_admin() and st.session_state["action_log"]:
        st.markdown("---")
        st.markdown(f"**{t('sidebar_recent_actions')}**")
        for entry in st.session_state["action_log"][-5:]:
            st.markdown(f"<small>{entry}</small>", unsafe_allow_html=True)

    # Sign out
    st.markdown("---")
    if st.button(t("sidebar_logout"), use_container_width=True):
        for k in ["authenticated", "role", "sp_path", "sp_name", "action_log"]:
            st.session_state.pop(k, None)
        st.rerun()


# ── Guards ────────────────────────────────────────────────────────────────────
def require_spreadsheet():
    if not sp():
        st.warning(t("no_spreadsheet_warning"))
        st.stop()

def require_admin():
    if not is_admin():
        st.warning("This page requires Secretariat access.")
        st.stop()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Dashboard  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
if page == "dashboard":
    require_admin()
    st.title(t("page_dashboard"))
    require_spreadsheet()

    wb      = _load_wb()
    df_cont = _sheet_to_df(wb, "CONTENT")
    df_mech = _sheet_to_df(wb, "ANNEXES")
    df_rq   = _sheet_to_df(wb, "REVIEW_QUEUE")
    today   = datetime.date.today()

    total_blocks = len(df_cont)
    over_limit   = int(df_cont["within_limit"].apply(
        lambda x: str(x or "").startswith("OVER")).sum()) if "within_limit" in df_cont else 0

    pending_rq = 0
    if not df_rq.empty and "status" in df_rq:
        pending_rq = int((df_rq["status"].str.upper() == "PENDING").sum())

    mechs_due = 0
    if not df_mech.empty and "next_verify_due" in df_mech:
        window = today + datetime.timedelta(days=14)
        for _, row in df_mech.iterrows():
            nvd = row.get("next_verify_due")
            if nvd:
                try:
                    if isinstance(nvd, datetime.datetime): nvd = nvd.date()
                    elif not isinstance(nvd, datetime.date): nvd = datetime.date.fromisoformat(str(nvd))
                    if nvd <= window: mechs_due += 1
                except (ValueError, TypeError):
                    pass
            if str(row.get("status", "")).upper() == "VERIFY": mechs_due += 1

    c1, c2, c3, c4 = st.columns(4)
    for col, num, label, colour in [
        (c1, total_blocks, t("dash_content_blocks"),  TEAL),
        (c2, over_limit,   t("dash_over_limit"),       PINK if over_limit else MINT),
        (c3, pending_rq,   t("dash_awaiting_review"),  PINK if pending_rq else MINT),
        (c4, mechs_due,    t("dash_mechanisms_due"),   LIME if mechs_due else MINT),
    ]:
        col.markdown(
            f'<div class="stat-box" style="border-top-color:{colour}">'
            f'<div class="stat-num">{num}</div>'
            f'<div class="stat-lbl">{label}</div></div>',
            unsafe_allow_html=True,
        )
    st.markdown("")

    st.subheader(t("dash_mechanisms"))
    if not df_mech.empty:
        show_cols = [c for c in ["mech_id","mechanism_name","category","status",
                                  "platform_eligible","last_verified","next_verify_due"]
                     if c in df_mech.columns]
        st.dataframe(df_mech[show_cols], use_container_width=True, hide_index=True)
    else:
        st.info("No ANNEXES sheet found.")

    if not df_rq.empty and "status" in df_rq:
        st.subheader(t("dash_review_queue"))
        sc = df_rq["status"].value_counts().reset_index()
        sc.columns = ["Status", "Count"]
        st.dataframe(sc, use_container_width=True, hide_index=True)

    # Member feedback summary
    if _MF_SHEET in wb.sheetnames:
        df_mf = _sheet_to_df(wb, _MF_SHEET)
        if not df_mf.empty:
            st.subheader("Member Feedback")
            new_count = int((df_mf.get("status", pd.Series()) == "New").sum()) if "status" in df_mf else len(df_mf)
            st.info(f"**{len(df_mf)}** feedback submission(s) total — **{new_count}** new. View them on the Give Feedback page.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Check Annexes  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "check_mechanisms":
    require_admin()
    st.title(t("page_check_mechanisms"))
    st.markdown(t("check_desc"))
    require_spreadsheet()

    wb      = _load_wb()
    df_mech = _sheet_to_df(wb, "ANNEXES")
    today   = datetime.date.today()
    window  = today + datetime.timedelta(days=14)

    due = []
    if not df_mech.empty:
        for _, row in df_mech.iterrows():
            nvd    = row.get("next_verify_due")
            is_due = False
            if nvd:
                try:
                    if isinstance(nvd, datetime.datetime): nvd = nvd.date()
                    elif not isinstance(nvd, datetime.date): nvd = datetime.date.fromisoformat(str(nvd))
                    if nvd <= window: is_due = True
                except (ValueError, TypeError):
                    pass
            if str(row.get("status", "")).upper() == "VERIFY": is_due = True
            if is_due: due.append(row)

    if not due:
        st.success(t("check_no_due"))
    else:
        st.info(f"**{len(due)} {t('check_due_label')}**")
        st.dataframe(pd.DataFrame(due)[["mech_id","mechanism_name","category","next_verify_due","status"]],
                     use_container_width=True, hide_index=True)
        st.markdown("---")
        st.subheader(t("check_run_verification"))
        api_key = st.text_input(
            t("check_api_key"), type="password",
            value=os.environ.get("ANTHROPIC_API_KEY", ""),
        )
        if st.button(t("check_run_btn"), type="primary", disabled=not api_key):
            sys.path.insert(0, str(Path(__file__).parent))
            try:
                import generate_toolkit as gt
            except ImportError as e:
                st.error(f"Could not import generate_toolkit: {e}"); st.stop()
            gt.SPREADSHEET = sp()
            progress = st.progress(0, text="Starting…")
            results  = []
            for i, (_, mrow) in enumerate(pd.DataFrame(due).iterrows()):
                mech_dict = mrow.to_dict()
                mid       = mech_dict.get("mech_id", "")
                mname     = mech_dict.get("mechanism_name", "")
                progress.progress(i / len(due), text=f"Checking {mid} — {mname}…")
                results.append((mech_dict, gt.call_ai_agent(mech_dict, api_key)))
            progress.progress(1.0, text="Done!")

            wb2   = openpyxl.load_workbook(sp())
            ws_m  = wb2["ANNEXES"]
            ws_rq = wb2["REVIEW_QUEUE"]
            mhdrs = [c.value for c in ws_m[2]]
            mcm   = {h: i for i, h in enumerate(mhdrs) if h}
            max_n = 0
            for row in ws_rq.iter_rows(min_row=3, values_only=True):
                rid = str(row[1] or "")
                if rid.startswith("RQ-"):
                    try: max_n = max(max_n, int(rid[3:]))
                    except ValueError: pass
            next_id = [max_n]
            mech_row_map = {}
            for r_idx, row in enumerate(ws_m.iter_rows(min_row=3, values_only=False), start=3):
                mid = row[0].value
                if mid: mech_row_map[str(mid).strip()] = r_idx

            summary = {"NO_CHANGE": 0, "CHANGE_DETECTED": 0, "UNABLE_TO_VERIFY": 0}
            for mech_dict, result in results:
                mid    = mech_dict.get("mech_id", "")
                name   = mech_dict.get("mechanism_name", "")
                cat    = mech_dict.get("category", "")
                status = result.get("status", "UNABLE_TO_VERIFY")
                conf   = result.get("confidence", "LOW")
                summary[status] = summary.get(status, 0) + 1
                if status == "NO_CHANGE":
                    row_idx = mech_row_map.get(str(mid).strip())
                    if row_idx: gt._update_mech_verified(ws_m, row_idx, mcm, today, cat)
                elif status == "CHANGE_DETECTED":
                    for ch in result.get("changes", []):
                        gt._rq_append(ws_rq, {"date_flagged": str(today), "mech_id": mid,
                            "mechanism_name": name, "category": cat, "change_type": "UPDATED_INFO",
                            "field": ch.get("field",""), "current_value": ch.get("current_value",""),
                            "proposed_value": ch.get("proposed_value",""), "reason": ch.get("reason",""),
                            "source_url": ch.get("source_url",""), "confidence": conf}, next_id)
                    for nm in result.get("new_mechanisms_found", []):
                        gt._rq_append(ws_rq, {"date_flagged": str(today), "mech_id": mid,
                            "mechanism_name": nm.get("name",""), "category": cat,
                            "change_type": "NEW_ENTRY", "field": "new_mechanism", "current_value": "",
                            "proposed_value": (f"{nm.get('name','')}  |  {nm.get('organisation','')}  |  {nm.get('url','')}"),
                            "reason": nm.get("reason",""), "source_url": nm.get("url",""), "confidence": conf}, next_id)
                else:
                    gt._rq_append(ws_rq, {"date_flagged": str(today), "mech_id": mid,
                        "mechanism_name": name, "category": cat, "change_type": "UNABLE_TO_VERIFY",
                        "field": "all", "current_value": "", "proposed_value": "",
                        "reason": result.get("notes",""), "source_url": "", "confidence": conf}, next_id)
            wb2.save(sp())
            st.success(t("check_done_msg").format(
                no_change=summary["NO_CHANGE"], changed=summary["CHANGE_DETECTED"], unable=summary["UNABLE_TO_VERIFY"]))
            if summary["CHANGE_DETECTED"] or summary["UNABLE_TO_VERIFY"]:
                st.info(t("check_review_hint"))
            st.session_state["action_log"].append(f"{today} — Checked {len(due)} annex entries")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Review Queue  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "review_queue":
    require_admin()
    st.title(t("page_review_queue"))
    require_spreadsheet()

    wb    = _load_wb(data_only=False)
    ws_rq = wb["REVIEW_QUEUE"] if "REVIEW_QUEUE" in wb.sheetnames else None
    if ws_rq is None:
        st.warning("No REVIEW_QUEUE sheet found."); st.stop()

    rq_hdrs = [c.value for c in ws_rq[2]]
    rqcm    = {h: i for i, h in enumerate(rq_hdrs) if h}

    col_f1, col_f2 = st.columns([2, 1])
    with col_f1:
        filter_status = st.multiselect(
            t("rq_filter_status"),
            ["PENDING","APPROVED","REJECTED","COMPLETED","COMPLETED_MANUAL"],
            default=["PENDING"],
        )
    with col_f2:
        filter_cat = st.multiselect(t("rq_filter_cat"), ["legal","emergency-funding","digital-security"])

    rows = []
    for r_idx, row in enumerate(ws_rq.iter_rows(min_row=3, values_only=False), start=3):
        rid    = row[rqcm.get("review_id", 1)].value
        if not rid: continue
        status = str(row[rqcm.get("status", 13)].value or "").strip().upper()
        cat    = str(row[rqcm.get("category", 5)].value or "")
        if filter_status and status not in [s.upper() for s in filter_status]: continue
        if filter_cat and cat not in filter_cat: continue
        rows.append((r_idx, row, rid, status, cat))

    if not rows:
        st.info(t("rq_no_items"))
    else:
        st.markdown(f"**{len(rows)} {t('rq_items_shown')}**"); st.markdown("")

    action_taken = False
    for r_idx, row, rid, status, cat in rows:
        mech   = str(row[rqcm.get("mechanism_name",  4)].value or "")
        ctyp   = str(row[rqcm.get("change_type",     6)].value or "")
        fld    = str(row[rqcm.get("field",            7)].value or "")
        cur    = str(row[rqcm.get("current_value",    8)].value or "")
        prop   = str(row[rqcm.get("proposed_value",   9)].value or "")
        reason = str(row[rqcm.get("reason",          10)].value or "")
        src    = str(row[rqcm.get("source_url",      11)].value or "")
        conf   = str(row[rqcm.get("confidence",      12)].value or "")
        date_f = str(row[rqcm.get("date_flagged",     2)].value or "")
        card_cls = {"PENDING":"forus-card change","APPROVED":"forus-card approved",
                    "REJECTED":"forus-card rejected"}.get(status, "forus-card")
        st.markdown(
            f'<div class="{card_cls}"><strong>{rid}</strong> &nbsp; {_badge(conf)} &nbsp; '
            f'<span style="color:#888;font-size:13px">{ctyp} · {cat} · {date_f}</span><br>'
            f'<strong>{mech}</strong> — <em>{fld}</em></div>',
            unsafe_allow_html=True,
        )
        ic1, ic2 = st.columns(2)
        with ic1:
            st.markdown(f"**{t('rq_current_value')}**")
            st.markdown(f"<div style='background:#FFF8F8;padding:8px;border-radius:4px;font-size:13px'>{cur or '—'}</div>", unsafe_allow_html=True)
        with ic2:
            st.markdown(f"**{t('rq_proposed_value')}**")
            st.markdown(f"<div style='background:#F8FFF8;padding:8px;border-radius:4px;font-size:13px'>{prop or '—'}</div>", unsafe_allow_html=True)
        if reason: st.markdown(f"<small>**Reason:** {reason}</small>", unsafe_allow_html=True)
        if src:    st.markdown(f"<small>**Source:** <a href='{src}' target='_blank'>{src}</a></small>", unsafe_allow_html=True)
        if status == "PENDING":
            bc1, bc2, _ = st.columns([1, 1, 4])
            with bc1:
                if st.button(t("rq_approve"), key=f"approve_{rid}"):
                    row[rqcm.get("status",13)].value       = "APPROVED"
                    row[rqcm.get("reviewed_by",15)].value  = "Forus staff"
                    row[rqcm.get("reviewed_date",16)].value= str(datetime.date.today())
                    wb.save(sp()); action_taken = True
                    st.session_state["action_log"].append(f"{datetime.date.today()} — Approved {rid}")
                    st.rerun()
            with bc2:
                if st.button(t("rq_reject"), key=f"reject_{rid}"):
                    row[rqcm.get("status",13)].value       = "REJECTED"
                    row[rqcm.get("reviewed_by",15)].value  = "Forus staff"
                    row[rqcm.get("reviewed_date",16)].value= str(datetime.date.today())
                    wb.save(sp()); action_taken = True
                    st.session_state["action_log"].append(f"{datetime.date.today()} — Rejected {rid}")
                    st.rerun()
        st.markdown("---")
    if action_taken:
        st.success(t("rq_saved_hint"))


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Apply Approved  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "apply_approved":
    require_admin()
    st.title(t("page_apply_approved"))
    st.markdown(t("apply_desc"))
    require_spreadsheet()

    wb    = _load_wb(data_only=False)
    ws_rq = wb["REVIEW_QUEUE"] if "REVIEW_QUEUE" in wb.sheetnames else None
    if ws_rq is None:
        st.warning("No REVIEW_QUEUE sheet found."); st.stop()

    rq_hdrs  = [c.value for c in ws_rq[2]]
    rqcm     = {h: i for i, h in enumerate(rq_hdrs) if h}
    approved = []
    for r_idx, row in enumerate(ws_rq.iter_rows(min_row=3, values_only=False), start=3):
        rid = row[rqcm.get("review_id", 1)].value
        if not rid: continue
        if str(row[rqcm.get("status", 13)].value or "").strip().upper() == "APPROVED":
            approved.append((r_idx, row))

    if not approved:
        st.info(t("apply_no_approved")); st.stop()
    st.success(t("apply_ready").format(n=len(approved)))
    reviewer = st.text_input(t("apply_reviewer_name"), value="Forus staff")
    if st.button(t("apply_btn"), type="primary"):
        sys.path.insert(0, str(Path(__file__).parent))
        try:
            import generate_toolkit as gt
            gt.SPREADSHEET = sp()
        except ImportError as e:
            st.error(f"Could not import generate_toolkit: {e}"); st.stop()
        gt.apply_approved(reviewer_name=reviewer or "Forus staff")
        st.success(t("apply_done").format(n=len(approved)))
        st.session_state["action_log"].append(
            f"{datetime.date.today()} — Applied {len(approved)} approved change(s)")
        st.info(t("apply_download_hint"))


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Generate PDF  (all roles)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "generate_pdf":
    st.title(t("page_generate_pdf"))
    require_spreadsheet()

    try:
        import reportlab
    except ImportError:
        st.error("**reportlab** is not installed. Run `pip install reportlab`."); st.stop()

    sys.path.insert(0, str(Path(__file__).parent))
    try:
        import generate_toolkit as gt
    except ImportError as e:
        st.error(f"Could not import generate_toolkit: {e}"); st.stop()
    gt.SPREADSHEET = sp()

    # Members only see the custom PDF tab
    if is_admin():
        tab1, tab2 = st.tabs([t("pdf_tab_standard"), t("pdf_tab_custom")])
    else:
        tab2 = st.container()
        tab1 = None  # not used for members

    if tab1 is not None:
        with tab1:
            st.subheader(t("pdf_tab_standard"))
            st.markdown(t("pdf_standard_desc"))
            col_a, col_b = st.columns(2)
            with col_a: build_public  = st.checkbox("Public PDF",  value=True)
            with col_b: build_network = st.checkbox("Network PDF (confidential)", value=True)
            if st.button(t("pdf_build_btn"), type="primary"):
                with tempfile.TemporaryDirectory() as tmp_dir:
                    v        = gt.VERSION
                    pub_path = os.path.join(tmp_dir, f"Forus_Toolkit_v{v}_Public.pdf")
                    net_path = os.path.join(tmp_dir, f"Forus_Toolkit_v{v}_Network.pdf")
                    gt.SPREADSHEET = sp(); gt.OUT_PUBLIC = pub_path; gt.OUT_NETWORK = net_path
                    with st.spinner("Building PDF(s)…"):
                        if build_public:  gt.build_pdf(1)
                        if build_network: gt.build_pdf(2)
                    dl_c1, dl_c2 = st.columns(2)
                    if build_public and os.path.exists(pub_path):
                        with open(pub_path, "rb") as f:
                            dl_c1.download_button(t("pdf_download_public"), f.read(),
                                file_name=f"Forus_Toolkit_v{v}_Public.pdf", mime="application/pdf")
                    if build_network and os.path.exists(net_path):
                        with open(net_path, "rb") as f:
                            dl_c2.download_button(t("pdf_download_network"), f.read(),
                                file_name=f"Forus_Toolkit_v{v}_Network.pdf", mime="application/pdf")
                    st.session_state["action_log"].append(f"{datetime.date.today()} — Generated PDF v{v}")

    with tab2:
        st.subheader(t("pdf_tab_custom"))
        st.markdown(t("pdf_custom_desc"))

        _PART_LABELS = {
            1: t("pdf_part_1"), 2: t("pdf_part_2"), 3: t("pdf_part_3"),
            4: t("pdf_part_4"), 5: t("pdf_part_5"), 6: t("pdf_part_6"), 7: t("pdf_part_7"),
        }
        _ANNEX_KEYS = {
            "Annex A: Legal Pro Bono Support":              t("pdf_annex_a"),
            "Annex B: Emergency Grants Mechanisms":         t("pdf_annex_b"),
            "Annex C: Physical & Digital Security Support": t("pdf_annex_c"),
        }
        _REGION_LABELS = [
            (t("pdf_region_africa"),  "africa"),
            (t("pdf_region_asia"),    "asia"),
            (t("pdf_region_europe"),  "europe"),
            (t("pdf_region_latam"),   "latam"),
            (t("pdf_region_pacific"), "pacific"),
            (t("pdf_region_global"),  "global"),
        ]
        _TOOL_LABELS = [
            ("T1", t("pdf_tool_t1")), ("T2", t("pdf_tool_t2")),
            ("T3", t("pdf_tool_t3")), ("T4", t("pdf_tool_t4")),
        ]
        _APPENDIX_TOOL_LABELS = [
            ("A1", t("pdf_appendix_a1")), ("A2", t("pdf_appendix_a2")), ("A3", t("pdf_appendix_a3")),
        ]
        _access_opts = [t("pdf_access_public_opt"), t("pdf_access_network_opt")]

        with st.form("custom_pdf_form"):
            c1, c2 = st.columns(2)
            with c1:
                cust_name  = st.text_input(t("pdf_contact_name"), placeholder="e.g. Maria Rodriguez")
                cust_org   = st.text_input(t("pdf_organisation"), placeholder="e.g. CCFD-Terre Solidaire")
            with c2:
                cust_email  = st.text_input(t("pdf_email"), placeholder="contact@example.org")
                cust_access = st.radio(t("pdf_access_level"), _access_opts, horizontal=True)

            st.markdown(t("pdf_parts_include"))
            pcols = st.columns(4)
            sel_parts = {}
            for i, (pnum, plabel) in enumerate(_PART_LABELS.items()):
                with pcols[i % 4]:
                    sel_parts[pnum] = st.checkbox(plabel, value=True)

            st.markdown(t("pdf_annexes"))
            acols = st.columns(3)
            sel_annexes = {}
            for i, (ann_key, ann_display) in enumerate(_ANNEX_KEYS.items()):
                with acols[i]:
                    sel_annexes[ann_key] = st.checkbox(ann_display, value=True)

            st.markdown(t("pdf_regions"))
            all_regions = st.checkbox(t("pdf_all_regions"), value=False)
            rcols = st.columns(6)
            sel_regions = {}
            for i, (rlabel, rkey) in enumerate(_REGION_LABELS):
                with rcols[i]:
                    sel_regions[rkey] = st.checkbox(rlabel, value=(rkey == "global"), disabled=all_regions)

            st.markdown(t("pdf_tools"))
            tcols = st.columns(2)
            sel_tools = {}
            for i, (tid, tlabel) in enumerate(_TOOL_LABELS):
                with tcols[i % 2]:
                    sel_tools[tid] = st.checkbox(tlabel, value=True, key=f"tool_{tid}")

            st.markdown(t("pdf_appendix_tools"))
            atcols = st.columns(3)
            for i, (tid, tlabel) in enumerate(_APPENDIX_TOOL_LABELS):
                with atcols[i]:
                    sel_tools[tid] = st.checkbox(tlabel, value=True, key=f"tool_{tid}")

            submitted = st.form_submit_button(t("pdf_build_custom_btn"), type="primary")

        if submitted:
            if not cust_name.strip() or not cust_org.strip():
                st.error(t("pdf_no_name_org_error"))
            else:
                effective_regions = ({rkey: True for _, rkey in _REGION_LABELS} if all_regions else sel_regions)
                req = {"req_id":"custom", "name":cust_name.strip(), "org":cust_org.strip(),
                       "email":cust_email.strip(), "parts":sel_parts, "regions":effective_regions,
                       "annexes":sel_annexes, "tools":sel_tools}
                access_level = 1 if _access_opts.index(cust_access) == 0 else 2
                v        = gt.VERSION
                safe_org = "".join(c if c.isalnum() or c in "-_" else "_" for c in cust_org.strip())
                fname    = f"Forus_Toolkit_v{v}_{'Public' if access_level==1 else 'Network'}_{safe_org}.pdf"

                with tempfile.TemporaryDirectory() as tmp_dir:
                    out_path = os.path.join(tmp_dir, fname)
                    gt.SPREADSHEET = sp()
                    gt.OUT_PUBLIC  = out_path if access_level == 1 else os.path.join(tmp_dir, "pub.pdf")
                    gt.OUT_NETWORK = out_path if access_level == 2 else os.path.join(tmp_dir, "net.pdf")
                    with st.spinner(f"Building custom PDF for {cust_name}…"):
                        try:
                            success = gt.build_pdf_from_request_dict(req, access_level=access_level, out_path=out_path)
                        except Exception as _e:
                            import traceback
                            st.error(f"Exception: {_e}"); st.code(traceback.format_exc()); success = False
                    if success and os.path.exists(out_path):
                        with open(out_path, "rb") as f:
                            st.download_button(f"⬇ Download PDF for {cust_org}", f.read(),
                                file_name=fname, mime="application/pdf")
                        if is_admin():
                            st.session_state["action_log"].append(
                                f"{datetime.date.today()} — Custom PDF for {cust_name}, {cust_org}")
                    else:
                        st.error(t("pdf_gen_failed"))


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Give Feedback  (all roles; admin sees extra review tab)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "give_feedback":
    st.title(t("page_give_feedback"))

    # Build part options list (general + 7 parts + 3 annexes)
    _fb_part_opts = [
        t("feedback_part_general"),
        t("pdf_part_1"), t("pdf_part_2"), t("pdf_part_3"), t("pdf_part_4"),
        t("pdf_part_5"), t("pdf_part_6"), t("pdf_part_7"),
        t("pdf_annex_a"), t("pdf_annex_b"), t("pdf_annex_c"),
    ]
    _fb_issue_opts = [t(f"feedback_issue_opt_{i}") for i in range(1, 6)]

    def _render_feedback_form():
        st.markdown(t("feedback_member_desc"))
        st.markdown("")
        with st.form("feedback_form", clear_on_submit=True):
            fc1, fc2 = st.columns(2)
            with fc1:
                fb_name = st.text_input(t("feedback_name_label"), placeholder="e.g. Maria Rodriguez")
            with fc2:
                fb_org  = st.text_input(t("feedback_org_label"),  placeholder="e.g. CCFD-Terre Solidaire")

            fb_part    = st.selectbox(t("feedback_part_label"),    _fb_part_opts)
            fb_section = st.text_input(t("feedback_section_label"), placeholder="e.g. Section 3.2 — compliance steps")
            fb_issue   = st.selectbox(t("feedback_issue_label"),   _fb_issue_opts)
            fb_desc    = st.text_area(t("feedback_desc_label"),    height=140,
                                      placeholder="Describe what seems wrong, outdated, or missing…")
            submitted  = st.form_submit_button(t("feedback_submit_btn"), type="primary")

        if submitted:
            if not fb_desc.strip():
                st.error(t("feedback_empty_error"))
                return
            data = {
                "timestamp":    str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
                "name":         fb_name.strip(),
                "organisation": fb_org.strip(),
                "language":     st.session_state.get("lang", "EN"),
                "part":         fb_part,
                "section":      fb_section.strip(),
                "issue_type":   fb_issue,
                "description":  fb_desc.strip(),
                "status":       "New",
            }
            # ── Feedback delivery — re-enable when SMTP secrets are configured ──
            # _save_feedback_to_spreadsheet(data)
            # ok, err_msg = _send_feedback_email(data)
            # if not ok:
            #     st.session_state["action_log"].append(f"⚠ Email send failed: {err_msg}")
            # ─────────────────────────────────────────────────────────────────
            st.success(t("feedback_success"))

    if is_admin():
        tab_submit, tab_view = st.tabs([
            t("feedback_admin_tab_submit"),
            t("feedback_admin_tab_view"),
        ])
        with tab_submit:
            _render_feedback_form()
        with tab_view:
            if not sp():
                st.info(t("no_spreadsheet_warning"))
            else:
                wb = _load_wb(data_only=False)
                if _MF_SHEET not in wb.sheetnames:
                    st.info(t("feedback_admin_no_entries"))
                else:
                    df_mf = _sheet_to_df(wb, _MF_SHEET)
                    if df_mf.empty:
                        st.info(t("feedback_admin_no_entries"))
                    else:
                        # Rename columns for display
                        col_map = {
                            "timestamp":    t("feedback_admin_col_time"),
                            "name":         t("feedback_admin_col_name"),
                            "organisation": t("feedback_admin_col_org"),
                            "part":         t("feedback_admin_col_part"),
                            "issue_type":   t("feedback_admin_col_issue"),
                            "description":  t("feedback_admin_col_desc"),
                            "status":       t("feedback_admin_col_status"),
                        }
                        show_cols = [c for c in col_map if c in df_mf.columns]
                        display   = df_mf[show_cols].rename(columns=col_map)
                        st.dataframe(display, use_container_width=True, hide_index=True)

                        # Mark-as-done buttons
                        st.markdown("---")
                        ws_mf = wb[_MF_SHEET]
                        mf_hdrs = [c.value for c in ws_mf[2]]
                        if "status" in mf_hdrs:
                            si = mf_hdrs.index("status") + 1
                            for idx, row_data in df_mf.iterrows():
                                ts  = row_data.get("timestamp", "")
                                sts = str(row_data.get("status", "New"))
                                if sts == "New":
                                    xr = idx + 3  # sheet row (2 header rows)
                                    if st.button(f"✓ Mark as reviewed — {ts}", key=f"mf_done_{idx}"):
                                        ws_mf.cell(xr, si).value = "Reviewed"
                                        wb.save(sp())
                                        if st.secrets.get("GDRIVE_CREDENTIALS"):
                                            save_to_gdrive()
                                        st.rerun()
    else:
        _render_feedback_form()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: Manage Tools  (admin only)
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "manage_tools":
    require_admin()
    st.title(t("page_manage_tools"))
    st.markdown(t("tools_desc"))
    require_spreadsheet()

    sys.path.insert(0, str(Path(__file__).parent))
    try:
        import generate_toolkit as gt
    except ImportError as e:
        st.error(f"Could not import generate_toolkit: {e}"); st.stop()
    gt.SPREADSHEET = sp()

    wb = _load_wb(data_only=False)
    if "TOOLS" not in wb.sheetnames:
        st.info("TOOLS sheet not found. Creating it with default content…")
        try:
            gt.ensure_tools_sheet(wb); wb.save(sp())
            st.success("✓ TOOLS sheet created with default content.")
            wb = _load_wb(data_only=False)
        except Exception as e:
            st.error(f"Could not create TOOLS sheet: {e}"); st.stop()

    ws_t   = wb["TOOLS"]
    t_hdrs = [c.value for c in ws_t[2]]
    try:
        ki = t_hdrs.index("field_key");  ti = t_hdrs.index("tool_id")
        ni = t_hdrs.index("tool_name");  li = t_hdrs.index("field_label")
        vi = t_hdrs.index("content_text"); fi = t_hdrs.index("change_flag")
    except ValueError as e:
        st.error(f"TOOLS sheet is missing expected columns: {e}"); st.stop()

    tools_rows = []; tools_row_idx = []
    for r_idx, row in enumerate(ws_t.iter_rows(min_row=3, values_only=False), start=3):
        fkey = row[ki].value
        if not fkey: continue
        tools_rows.append({"tool_id": str(row[ti].value or ""), "tool_name": str(row[ni].value or ""),
                            "field_key": str(fkey), "field_label": str(row[li].value or ""),
                            "content_text": str(row[vi].value or ""), "change_flag": str(row[fi].value or "OK")})
        tools_row_idx.append(r_idx)

    if not tools_rows:
        st.warning("No rows found in TOOLS sheet."); st.stop()

    tool_ids     = sorted(set(r["tool_id"] for r in tools_rows))
    sel_tool     = st.selectbox(t("tools_filter"), ["All"] + tool_ids)
    show_flagged = st.checkbox(t("tools_show_flagged"), value=False)
    filtered     = [r for r in tools_rows
                    if (sel_tool == "All" or r["tool_id"] == sel_tool)
                    and (not show_flagged or r["change_flag"] in ("VERIFY","UPDATED"))]
    st.markdown(f"**{len(filtered)} {t('tools_fields_shown')}**"); st.markdown("")

    for row_data, r_idx in [(r, tools_row_idx[tools_rows.index(r)]) for r in filtered]:
        flag        = row_data["change_flag"]
        flag_colour = {"VERIFY":"🟡","UPDATED":"🟢","OK":"⚪"}.get(flag, "⚪")
        with st.expander(f"{flag_colour} **{row_data['tool_id']}** · {row_data['field_label']}",
                         expanded=(flag in ("VERIFY","UPDATED"))):
            new_text = st.text_area(t("tools_content_label"), value=row_data["content_text"],
                                    height=100, key=f"tools_text_{r_idx}")
            cf1, cf2 = st.columns([2, 1])
            with cf1:
                new_flag = st.selectbox(t("tools_status"), ["OK","VERIFY","UPDATED"],
                    index=["OK","VERIFY","UPDATED"].index(flag) if flag in ["OK","VERIFY","UPDATED"] else 0,
                    key=f"tools_flag_{r_idx}")
            with cf2:
                if st.button(t("tools_save_btn"), key=f"tools_save_{r_idx}"):
                    ws_t.cell(r_idx, vi+1).value = new_text
                    ws_t.cell(r_idx, fi+1).value = new_flag
                    ws_t.cell(r_idx, t_hdrs.index("last_updated")+1).value = str(datetime.date.today())
                    wb.save(sp())
                    st.success(f"✓ Saved {row_data['field_key']}")
                    st.rerun()

    st.markdown("---")
    st.markdown(t("tools_instructions"))
